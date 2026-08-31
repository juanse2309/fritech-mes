from flask import Blueprint, request, session, current_app
from sqlalchemy import text
from io import BytesIO
import os
import tempfile
from backend.core.responses import api_success, api_error
from backend.core import task_runner
from backend.utils.auth_middleware import require_role, ROL_ADMINS, _obtener_usuario_activo
from backend.models.sql_models import db, ProduccionPulido, PncInyeccion, PncPulido, PncEnsamble, BujeRevuelto, Producto, TrazabilidadLote, PulidoOverride
from backend.utils.formatters import normalizar_codigo, preservar_o_normalizar_prefijo, normalizar_codigo_sin_prefijo
from backend.services.audit_service import AuditService, OwnershipMismatchException, TurnoInvalidoException
from backend.services.pulido_service import PulidoService, FechaPulidoInvalidaException, CantidadExcedeInyectadoException
from backend.services.pausas_service import PausasService
from backend.utils.time_utils import get_colombia_time
import uuid
from datetime import datetime
import logging
import json

logger = logging.getLogger(__name__)
pulido_bp = Blueprint('pulido_bp', __name__)

ROLES_PULIDO = ROL_ADMINS + ['JEFE PULIDO', 'PULIDO', 'JEFE AUXILIAR INVENTARIO']
# Panel de admin (ver/pausar/reanudar sesiones de TODAS las operarias, plan
# 2026-08-31) -- deliberadamente MÁS estrecho que ROLES_PULIDO: ver y tocar
# la sesión de otra persona es un privilegio de supervisión, no algo que
# cualquiera con acceso normal a Pulido deba poder hacer.
ROLES_ADMIN_PULIDO = ROL_ADMINS + ['JEFE PULIDO']


def _es_prueba_pulido(orden_produccion, id_pulido):
    """Sandbox de pruebas (9999/PRUEBA): ver es_prueba más abajo -- factorizado
    para poder evaluarlo también contra el estado ANTERIOR de un registro al
    revertir su efecto en inventario (ver _ejecutar_persistencia_pulido)."""
    return (
        '9999' in str(orden_produccion or '').upper() or 'PRUEBA' in str(orden_produccion or '').upper()
        or '9999' in str(id_pulido or '').upper() or 'PRUEBA' in str(id_pulido or '').upper()
    )


def _ejecutar_persistencia_pulido(registro, data, responsable, ahora,
                                   forzar_bloqueo=False, motivo_forzado=None, autorizado_por=None):
    """
    Función privada auxiliar para encapsular la persistencia y la lógica de negocio
    compleja del módulo de pulido. Mantiene la ruta compacta (< 50 líneas).

    forzar_bloqueo/motivo_forzado/autorizado_por: bypass de los bloqueos duros
    de fecha y cantidad (plan 2026-08-28) -- el guard de que solo un ADMIN
    puede forzar vive en la ruta (registrar_pulido), no aquí. Cuando se
    fuerza y algún bloqueo SÍ se habría disparado, queda una fila en
    PulidoOverride (el "reporte para restar puntos" que pidió la jefa).
    """
    id_pulido = data.get('id_pulido')

    # Snapshot del efecto en inventario YA APLICADO por este registro, ANTES
    # de sobreescribir sus campos -- necesario para revertirlo abajo. Sin
    # esto, editar un reporte ya guardado (o que la tablet lo reenvíe por un
    # reintento de red) vuelve a descontar por_pulir y a sumar p_terminado
    # una segunda vez para la MISMA producción física (hallazgo 2026-08-27,
    # el "flujo directo" no era idempotente: solo Inyección tenía el guard
    # `estado == 'CERRADO'`, Pulido nunca lo tuvo).
    efecto_anterior = None
    if registro:
        rev_anterior = sum(
            float(r.cantidad or 0) for r in
            db.session.query(BujeRevuelto).filter_by(id_pulido=registro.id_pulido).all()
        )
        efecto_anterior = {
            'codigo': registro.codigo,
            'buenas': float(registro.cantidad_real or 0),
            'total': (
                float(registro.cantidad_real or 0) + float(registro.pnc_inyeccion or 0)
                + float(registro.pnc_pulido or 0) + rev_anterior
            ),
            'es_prueba': _es_prueba_pulido(registro.orden_produccion, registro.id_pulido),
        }

    if not registro:
        # Si no existe (o fue borrado de la DB), crear uno nuevo para evitar Error 500
        if id_pulido:
            logger.warning(f" [RECOVERY] id_pulido {id_pulido} no encontrado en DB. Creando nuevo registro.")
        registro = ProduccionPulido(id_pulido=id_pulido or f"PUL-{ahora.strftime('%Y%m%d%H%M%S')}", fecha_registro=ahora)
        db.session.add(registro)

    if not getattr(registro, 'fecha_registro', None):
        registro.fecha_registro = ahora

    # Mapeo y Estandarización
    registro.fecha = datetime.strptime(data.get('fecha_inicio', ahora.strftime('%Y-%m-%d')), '%Y-%m-%d').date()
    # ── Blindaje: se preserva el prefijo que traiga (MT-, CAR-, FR-) y los
    #    números puros quedan intactos; nunca se les antepone una división ──
    registro.codigo = preservar_o_normalizar_prefijo(data.get('codigo_producto'))
    registro.responsable = responsable
    registro.cantidad_real = float(data.get('cantidad_real') or 0)
    registro.pnc_inyeccion = int(data.get('pnc_inyeccion') or 0)
    registro.pnc_pulido = int(data.get('pnc_pulido') or 0)
    registro.criterio_pnc_inyeccion = data.get('criterio_pnc_inyeccion')
    registro.criterio_pnc_pulido = data.get('criterio_pnc_pulido')
    registro.orden_produccion = data.get('orden_produccion') or 'SIN OP'
    registro.observaciones = data.get('observaciones', '')
    registro.estado = data.get('estado', 'FINALIZADO')
    registro.departamento = 'Pulido'  # Estandarización exigida
    registro.lote = data.get('lote') or 'SIN LOTE'
    registro.cantidad_recibida = float(data.get('cantidad_recibida') or 0)
    registro.almacen_destino = data.get('almacen_destino', 'P. TERMINADO')

    # Validación de consistencia (Bujes Buenos + PNC <= Total)
    total_reportado = registro.cantidad_real + registro.pnc_inyeccion + registro.pnc_pulido
    if registro.cantidad_recibida < total_reportado:
        logger.warning(f" [VALIDATION] Inconsistencia en {registro.id_pulido}: Total {registro.cantidad_recibida} < Suma {total_reportado}")

    # ── Bloqueos duros (plan 2026-08-28): fecha same-day + cantidad <= inyectado ──
    # Se evalúan SIN forzar primero para saber si el forzado realmente saltó
    # algo (y así no dejar filas de override "fantasma" cuando forzar_bloqueo
    # viene en true pero nada se habría bloqueado). Si no se fuerza, la
    # excepción sube tal cual y _ejecutar_persistencia_pulido no persiste nada
    # (todavía no hubo commit).
    overrides_aplicados = []
    try:
        PulidoService.validar_bloqueo_fecha(registro.fecha, forzado=False)
    except FechaPulidoInvalidaException as exc_fecha:
        if not forzar_bloqueo:
            raise
        overrides_aplicados.append(('FECHA', str(exc_fecha)))

    try:
        PulidoService.validar_saldo_op(
            registro.orden_produccion, registro.codigo, total_reportado,
            id_pulido_actual=registro.id_pulido, forzado=False
        )
    except CantidadExcedeInyectadoException as exc_cant:
        if not forzar_bloqueo:
            raise
        overrides_aplicados.append(('CANTIDAD', str(exc_cant)))

    # Manejo de Horas y Cálculos de Tiempo
    if data.get('hora_inicio'):
        h_h, h_m = data['hora_inicio'].split(':')
        dt_ini = ahora.replace(hour=int(h_h), minute=int(h_m), second=0, microsecond=0)
        registro.hora_inicio = dt_ini.replace(tzinfo=None)
    
    if data.get('hora_fin'):
        h_h, h_m = data['hora_fin'].split(':')
        dt_fin = ahora.replace(hour=int(h_h), minute=int(h_m), second=0, microsecond=0)
        registro.hora_fin = dt_fin.replace(tzinfo=None)

    # Cálculo de Métricas
    if data.get('hora_inicio') and data.get('hora_fin'):
        hi_h, hi_m = data['hora_inicio'].split(':')
        hf_h, hf_m = data['hora_fin'].split(':')
        
        t_ini = ahora.replace(hour=int(hi_h), minute=int(hi_m), second=0, microsecond=0)
        t_fin = ahora.replace(hour=int(hf_h), minute=int(hf_m), second=0, microsecond=0)
        
        diff = t_fin - t_ini
        segundos_segmento = int(diff.total_seconds())
        if segundos_segmento < 0: segundos_segmento += 86400

        # Barrera arquitectónica: rechaza duraciones imposibles (típico error de
        # digitar 3:20 en vez de 13:20) antes de persistir nada.
        PulidoService.validar_duracion_turno(segundos_segmento)

        tiempo_acumulado_ms = float(data.get('tiempo_acumulado_ms') or 0)
        segundos_totales = segundos_segmento + int(tiempo_acumulado_ms / 1000)

        # Autoridad matemática del descuento por pausas programadas: PausasService.
        # El frontend solo aporta hora_inicio/hora_fin crudas (Zero Trust) — el
        # cálculo de intersección de intervalos vive exclusivamente en la capa de servicio.
        descuento_info = PausasService.calcular_descuento_pausas_programadas(t_ini, t_fin)
        segundos_descuento = descuento_info['segundos_descuento']
        if segundos_descuento > 0:
            segundos_totales = max(0, segundos_totales - segundos_descuento)

        registro.duracion_segundos = segundos_totales
        registro.tiempo_total_minutos = float(round(segundos_totales / 60.0, 2))

        cant = float(registro.cantidad_real or 0)
        if cant > 0:
            registro.segundos_por_unidad = float(round(segundos_totales / cant, 2))
        else:
            registro.segundos_por_unidad = 0.0

        if descuento_info['detalle']:
            payload = {
                "descuento_programado_min": round(segundos_descuento / 60.0, 2),
                "detalle": descuento_info['detalle']
            }
            tag = f"[AUTO_BREAK]{json.dumps(payload, ensure_ascii=False)}[/AUTO_BREAK]"
            obs = (registro.observaciones or "")
            if "[AUTO_BREAK]" in obs and "[/AUTO_BREAK]" in obs:
                pre = obs.split("[AUTO_BREAK]")[0]
                post = obs.split("[/AUTO_BREAK]")[-1]
                registro.observaciones = (pre + tag + post).strip()
            else:
                registro.observaciones = (obs + "\n" + tag).strip() if obs else tag

        logger.debug(f" [TIME-DEBUG] {registro.id_pulido} -> Seg: {segundos_segmento}s, Acum: {tiempo_acumulado_ms}ms, DescProgramado: {segundos_descuento}s, Total: {segundos_totales}s")
    else:
        registro.duracion_segundos = 0
        registro.tiempo_total_minutos = 0.0
        registro.segundos_por_unidad = 0.0

    db.session.flush()
 
    # Sincronización de PNC Detallado
    # registro.codigo puede llegar con o sin prefijo 'FR-' (según lo reportó la
    # planta), pero las tablas de PNC se indexan sin prefijo para no fragmentar
    # 'FR-1005' / '1005' en filas distintas — se sanitiza aquí explícitamente
    # antes de tocar el ORM. Los prefijos de otras divisiones (MT-, CAR-) se
    # conservan: normalizar_codigo_sin_prefijo() solo quita 'FR-'.
    codigo_pnc = normalizar_codigo_sin_prefijo(registro.codigo)
    # Operaria a la que se atribuye toda merma de pulido de este turno. Se
    # resuelve una sola vez: si no hay persona identificable, el reporte no
    # debe persistirse con la merma huérfana.
    operaria_responsable = PulidoService.resolver_operaria_responsable(registro)
    # Dueño de la merma de INYECCIÓN detectada en pulido: el operario que fabricó
    # las piezas, rastreado por el lote/OP de origen. Puede ser None si el lote
    # padre no es rastreable — en ese caso la fila queda sin atribuir a propósito.
    operario_inyeccion_origen = PulidoService.resolver_operario_inyeccion_origen(registro)
    pnc_detail = data.get('pnc_detail', [])
    db.session.query(PncInyeccion).filter_by(id_inyeccion=registro.id_pulido).delete()
    db.session.query(PncPulido).filter_by(id_pulido=registro.id_pulido).delete()
    db.session.query(PncEnsamble).filter_by(id_ensamble=registro.id_pulido).delete()

    if pnc_detail:
        for pnc_item in pnc_detail:
            proc = pnc_item.get('proceso', '').upper()
            cant = float(pnc_item.get('cantidad') or 0)
            crit = pnc_item.get('criterio', '')
            if cant <= 0: continue

            if proc == 'INYECCION':
                db.session.add(PncInyeccion(
                    id_pnc_inyeccion=uuid.uuid4().hex[:8],
                    id_inyeccion=registro.id_pulido,
                    id_codigo=codigo_pnc,
                    cantidad=cant,
                    criterio=crit,
                    responsable=operario_inyeccion_origen
                ))
            elif proc == 'PULIDO':
                db.session.add(PncPulido(
                    id_pnc_pulido=uuid.uuid4().hex[:8],
                    id_pulido=registro.id_pulido,
                    codigo=codigo_pnc,
                    cantidad=cant,
                    criterio=crit,
                    responsable=operaria_responsable
                ))
            elif proc == 'ENSAMBLE':
                db.session.add(PncEnsamble(
                    id_pnc_ensamble=uuid.uuid4().hex[:8],
                    id_ensamble=registro.id_pulido,
                    id_codigo=codigo_pnc,
                    cantidad=cant,
                    criterio=crit
                ))
    else:
        # Fix de Sincronización: si el payload trae el agregado (pnc_pulido/
        # pnc_inyeccion) SIN el desglose itemizado, la tabla hija no puede
        # quedar huérfana de la maestra — o el Dashboard (que lee de
        # db_pnc_pulido/db_pnc_inyeccion) subcuenta este PNC en silencio.
        # Se reutiliza el criterio de texto libre del header si vino, y si
        # no, se cae a un genérico explícito para no perder trazabilidad.
        if registro.pnc_pulido and registro.pnc_pulido > 0:
            db.session.add(PncPulido(
                id_pnc_pulido=uuid.uuid4().hex[:8],
                id_pulido=registro.id_pulido,
                codigo=codigo_pnc,
                cantidad=registro.pnc_pulido,
                criterio=registro.criterio_pnc_pulido or 'Diferencia/Sobrante (Sin Desglose)',
                responsable=operaria_responsable
            ))
        if registro.pnc_inyeccion and registro.pnc_inyeccion > 0:
            db.session.add(PncInyeccion(
                id_pnc_inyeccion=uuid.uuid4().hex[:8],
                id_inyeccion=registro.id_pulido,
                id_codigo=codigo_pnc,
                cantidad=registro.pnc_inyeccion,
                criterio=registro.criterio_pnc_inyeccion or 'Diferencia/Sobrante (Sin Desglose)',
                responsable=operario_inyeccion_origen
            ))

    db.session.flush()

    # Manejo de Bujes Revueltos
    revueltos_list = data.get('revueltos', [])
    logger.debug(f" [REVUELTOS-DEBUG] Payload recibido: {revueltos_list}")
    
    db.session.query(BujeRevuelto).filter_by(id_pulido=registro.id_pulido).delete()

    for rev_item in revueltos_list:
        cod_rev = preservar_o_normalizar_prefijo(rev_item.get('id_codigo'))
        cant_rev = float(rev_item.get('cantidad') or 0)
        
        if cant_rev <= 0 or not cod_rev: 
            continue

        db.session.add(BujeRevuelto(
            id_bujes_revueltos=uuid.uuid4().hex[:8],
            id_pulido=registro.id_pulido,
            id_codigo=cod_rev,
            cantidad=cant_rev,
            codigo_ensamble=cod_rev,
            responsable=registro.responsable
        ))

    # Distribución FIFO OP/Pedidos
    op_actual = registro.orden_produccion
    if op_actual and str(op_actual).strip() != 'SIN OP':
        from backend.models.sql_models import DistribucionOpPedidos
        op_limpia = str(registro.orden_produccion or '').strip()
        codigo_limpio = normalizar_codigo(registro.codigo)
        
        cubetas = db.session.query(DistribucionOpPedidos).filter(
            DistribucionOpPedidos.op_world_office == op_limpia,
            DistribucionOpPedidos.codigo_producto == codigo_limpio
        ).order_by(DistribucionOpPedidos.id_distribucion.asc()).all()

        piezas_por_repartir = float(registro.cantidad_real or 0)
        
        if not cubetas and piezas_por_repartir > 0:
            pedido_asoc = db.session.query(DistribucionOpPedidos.id_pedido).filter(
                DistribucionOpPedidos.op_world_office == op_limpia
            ).first()
            id_pedido_final = pedido_asoc[0] if (pedido_asoc and pedido_asoc[0]) else f"PED-IMPREVISTO-{op_limpia}"
            
            nueva_cubeta = DistribucionOpPedidos(
                op_world_office=op_limpia,
                id_pedido=id_pedido_final,
                codigo_producto=codigo_limpio,
                cant_requerida=piezas_por_repartir,
                cant_inyectada=piezas_por_repartir,
                cant_pulida=piezas_por_repartir,
                cant_ensamblada=0,
                cant_alistada=0
            )
            db.session.add(nueva_cubeta)
            db.session.flush()
            cubetas = [nueva_cubeta]
            piezas_por_repartir = 0.0
        
        for cubeta in cubetas:
            if piezas_por_repartir <= 0:
                break
            falta = max(0, (cubeta.cant_requerida or 0) - (cubeta.cant_pulida or 0))
            if falta > 0:
                if piezas_por_repartir >= falta:
                    cubeta.cant_pulida = (cubeta.cant_pulida or 0) + falta
                    piezas_por_repartir -= falta
                else:
                    cubeta.cant_pulida = (cubeta.cant_pulida or 0) + piezas_por_repartir
                    piezas_por_repartir = 0

    # Flujo directo: Actualizar inventario final en db_productos sin usar TrazabilidadLote
    try:
        buenas = float(registro.cantidad_real or 0)
        pnc_total = float(registro.pnc_inyeccion or 0) + float(registro.pnc_pulido or 0)
        rev_total = sum(float(r.get('cantidad', 0)) for r in revueltos_list)
        total_descuento_por_pulir = buenas + pnc_total + rev_total

        es_prueba = _es_prueba_pulido(registro.orden_produccion, registro.id_pulido)

        # Revertir el efecto que este MISMO registro ya haya aplicado antes
        # de aplicar el nuevo -- así una edición o un reenvío queda neto en
        # el delta real, en vez de descontar por_pulir dos veces (ver
        # snapshot al inicio de la función). Si el código cambió entre
        # ediciones, esto también corrige el producto correcto: revierte
        # sobre el código viejo y aplica sobre el nuevo.
        if efecto_anterior and not efecto_anterior['es_prueba']:
            codigo_ant = preservar_o_normalizar_prefijo(efecto_anterior['codigo'])
            # Match por codigo_sistema O id_codigo (no solo codigo_sistema):
            # el frontend de Pulido (normalizarCodigo) le quita el prefijo
            # 'FR-' antes de enviar -- preservar_o_normalizar_prefijo NO se
            # lo vuelve a poner a propósito (no debe inventar división), así
            # que un código FR- puro llega aquí como '9308', no 'FR-9308'.
            # db_productos.codigo_sistema SIEMPRE lleva el prefijo real, pero
            # id_codigo guarda la referencia tal cual -- sin este OR, CADA
            # reporte de Pulido de una referencia FR- (la división por
            # defecto, la mayoría del catálogo) no encontraba el producto y
            # el descuento de inventario se saltaba en silencio (hallazgo
            # 2026-08-31, probado en vivo en ambos modos Satélite y PRO).
            prod_anterior = db.session.query(Producto).filter(
                (Producto.codigo_sistema == codigo_ant) | (Producto.id_codigo == codigo_ant)
            ).first()
            if prod_anterior:
                prod_anterior.por_pulir = float(prod_anterior.por_pulir or 0) + efecto_anterior['total']
                p_terminado_revertido = float(prod_anterior.p_terminado or 0) - efecto_anterior['buenas']
                if p_terminado_revertido < 0:
                    logger.warning(
                        f"⚠️ [PULIDO-INVENTARIO] Al editar {registro.id_pulido}, revertir P.Terminado de "
                        f"{efecto_anterior['codigo']} lo manda a negativo ({p_terminado_revertido}) -- "
                        f"probablemente ya se consumió aguas abajo (empaque/despacho). Se deja en 0."
                    )
                prod_anterior.p_terminado = max(0, p_terminado_revertido)

        if not es_prueba and total_descuento_por_pulir > 0:
            codigo_actual = preservar_o_normalizar_prefijo(registro.codigo)
            prod_wip = db.session.query(Producto).filter(
                (Producto.codigo_sistema == codigo_actual) | (Producto.id_codigo == codigo_actual)
            ).first()
            if prod_wip:
                # Restar de por_pulir, evitando negativos de forma preventiva
                prod_wip.por_pulir = max(0, float(prod_wip.por_pulir or 0) - total_descuento_por_pulir)
                # Sumar las buenas a p_terminado
                prod_wip.p_terminado = float(prod_wip.p_terminado or 0) + buenas
        else:
            logger.debug(f"🧪 [SANDBOX] Lote de prueba {registro.id_pulido}. Se ignoró impacto en inventario.")
    except Exception as err:
        logger.error(f"Error actualizando inventario directo en pulido: {err}")

    for tipo_bloqueo, detalle_bloqueo in overrides_aplicados:
        db.session.add(PulidoOverride(
            id_pulido=registro.id_pulido,
            tipo=tipo_bloqueo,
            operaria=responsable,
            autorizado_por=autorizado_por,
            motivo=motivo_forzado,
            detalle=detalle_bloqueo,
        ))
        logger.warning(
            f"⚠️ [PULIDO-OVERRIDE] Bloqueo {tipo_bloqueo} de {registro.id_pulido} "
            f"forzado por {autorizado_por!r} (operaria {responsable!r}). Motivo: {motivo_forzado!r}."
        )

    db.session.commit()
    return {
        "success": True,
        "message": "Registro de pulido sincronizado (Flujo directo)",
        "id_pulido": registro.id_pulido,
        "upsert": "UPDATE" if id_pulido and registro.id else "INSERT",
        "overrides_aplicados": [t for t, _ in overrides_aplicados],
    }

@pulido_bp.route('/api/pulido', methods=['POST'])
@require_role(ROLES_PULIDO)
def registrar_pulido():
    try:
        data = request.get_json()
        if not data:
            return api_error("No data provided", status_code=400)

        ahora = get_colombia_time()

        # Lógica de Upsert (Evitar duplicados por id_pulido)
        id_pulido = data.get('id_pulido')
        registro = ProduccionPulido.query.filter_by(id_pulido=id_pulido).first() if id_pulido else None

        # Guard de ownership centralizado con AuditService e identificación desacoplada HTTP
        usuario_activo = _obtener_usuario_activo()
        candidato_responsable = data.get('responsable') or usuario_activo

        if not candidato_responsable or str(candidato_responsable).strip().upper() in ['', 'SISTEMA']:
            return api_error(
                "Se requiere una identidad de operario o responsable válida para registrar el proceso de pulido",
                status_code=400, code="RESPONSABLE_REQUERIDO"
            )

        try:
            responsable = AuditService.resolver_y_validar_propietario(registro, candidato_responsable)
        except OwnershipMismatchException as e:
            return api_error(
                e.message, status_code=409, code="PULIDO_SESSION_OWNERSHIP_MISMATCH",
                id_pulido=id_pulido, responsable_db=e.responsable_db, responsable_in=e.responsable_in
            )

        # Bloqueos duros de fecha/cantidad (plan 2026-08-28): solo un ADMIN
        # puede saltarlos, y solo con un motivo -- mismo patrón que
        # EnsambleService.cerrar_jornada(forzar=...).
        forzar_bloqueo = bool(data.get('forzar_bloqueo'))
        motivo_forzado = (data.get('motivo_forzado') or '').strip() or None
        rol_activo = session.get('role')
        if forzar_bloqueo:
            if not any(r in (rol_activo or '').upper() for r in ROL_ADMINS):
                return api_error(
                    "Solo un ADMIN puede autorizar el guardado de un reporte de Pulido bloqueado",
                    status_code=403, code="FORZAR_NO_AUTORIZADO"
                )
            if not motivo_forzado:
                return api_error(
                    "Forzar el guardado de un reporte bloqueado requiere indicar un motivo",
                    status_code=400, code="MOTIVO_REQUERIDO"
                )

        # Delegar la persistencia compleja a la función privada
        res = _ejecutar_persistencia_pulido(
            registro, data, responsable, ahora,
            forzar_bloqueo=forzar_bloqueo, motivo_forzado=motivo_forzado, autorizado_por=usuario_activo,
        )
        return api_success(data=res, status_code=201)

    except FechaPulidoInvalidaException as e:
        db.session.rollback()
        return api_error(
            e.message, status_code=409, code="PULIDO_FECHA_BLOQUEADA",
            fecha_reporte=str(e.fecha_reporte), hoy=str(e.hoy)
        )

    except CantidadExcedeInyectadoException as e:
        db.session.rollback()
        return api_error(
            e.message, status_code=409, code="PULIDO_CANTIDAD_EXCEDE_INYECTADO",
            orden_produccion=e.op, referencia=e.referencia,
            inyectado=e.inyectado, ya_reportado=e.ya_reportado, disponible=e.disponible
        )

    except TurnoInvalidoException as e:
        db.session.rollback()
        return api_error(e.message, status_code=400, code="TURNO_DURACION_INVALIDA")

    except ValueError as e:
        # Incluye la falta de operaria responsable para atribuir la merma
        # (PulidoService.resolver_operaria_responsable): es un error del
        # cliente, no un fallo interno.
        db.session.rollback()
        return api_error(str(e), status_code=400, code="RESPONSABLE_REQUERIDO")

    except Exception as e:
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"❌ [PULIDO-ERROR] Error crítico al registrar: {str(e)}\n{error_trace}")
        return api_error(
            str(e), status_code=500,
            detail="Error interno al procesar el reporte de pulido (Revisa logs del servidor)"
        )



@pulido_bp.route('/api/pulido/ultimo_registro', methods=['GET'])
def get_ultimo_registro_pulido():
    """
    Devuelve el último registro de db_pulido insertado por el operario
    indicado en el query param ?responsable=<nombre>.
    Responde con: { success, registro: { fecha_hora, codigo_producto, cantidad } | null }
    LIMIT 1 + ORDER BY id DESC → coste mínimo, usa el índice de id.
    """
    responsable = request.args.get('responsable', '').strip()
    if not responsable:
        return api_error("Falta parámetro 'responsable'", status_code=400)

    try:
        registro = (
            ProduccionPulido.query
            .filter(ProduccionPulido.responsable == responsable)
            .order_by(ProduccionPulido.id.desc())
            .first()
        )

        if not registro:
            return api_success(data={"registro": None})

        # Construir la fecha_hora combinando fecha + hora_fin (o hora_inicio como fallback)
        hora_ref = registro.hora_fin or registro.hora_inicio
        if registro.fecha and hora_ref:
            try:
                fecha_hora_str = f"{registro.fecha.strftime('%d/%m/%Y')} {hora_ref.strftime('%H:%M')}"
            except Exception:
                fecha_hora_str = str(registro.fecha)
        else:
            fecha_hora_str = registro.fecha.strftime('%d/%m/%Y') if registro.fecha else '—'

        return api_success(data={
            "registro": {
                "fecha_hora": fecha_hora_str,
                "codigo_producto": registro.codigo or '—',
                "cantidad": float(registro.cantidad_real or 0),
                "cantidad_aprobada": float(registro.cantidad_real or 0),
                "piezas": float(registro.cantidad_real or 0)
            }
        })

    except Exception as e:
        logger.error(f"[ultimo_registro] Error: {e}")
        return api_error(str(e), status_code=500)


@pulido_bp.route('/api/pulido/session_active', methods=['GET'])
def get_active_pulido_session():
    try:
        # Forzar limpieza de caché de sesión para obtener datos frescos de la DB
        db.session.remove()
        
        if request.args.get('ping') == 'true':
            return api_success(data={"ping": "pong"})

        responsable = request.args.get('responsable')
        if not responsable:
            return api_error("Falta responsable", status_code=400)

        # TTL Garbage Collector: autocierra sesiones zombi (>14h en TRABAJANDO/
        # EN_PROCESO/PAUSADO) antes de evaluar si hay una sesión activa, para
        # que un turno olvidado no bloquee indefinidamente el formulario.
        PulidoService.limpiar_sesiones_zombis(responsable)

        # FIX Efecto Daniela: Filtrar por estado EN_PROCESO/PAUSADO,
        # NO por hora_fin=None (registros migrados no tienen hora_fin)
        query = ProduccionPulido.query.filter(ProduccionPulido.responsable == responsable)
        
        id_especifico = request.args.get('id_pulido')
        if id_especifico:
            query = query.filter(ProduccionPulido.id_pulido == id_especifico)
        else:
            query = query.filter(ProduccionPulido.estado.in_(['EN_PROCESO', 'PAUSADO', 'PAUSADO_COLA', 'TRABAJANDO']))
            
        sesion = query.order_by(ProduccionPulido.id.desc()).first()

        if sesion:
            return api_success(data={
                "session": {
                    "id_pulido": sesion.id_pulido,
                    "codigo": sesion.codigo,
                    "lote": sesion.lote,
                    "orden_produccion": sesion.orden_produccion,
                    "hora_inicio_dt": sesion.hora_inicio.isoformat() if sesion.hora_inicio else None,
                    "duracion_segundos": sesion.duracion_segundos or 0,
                    "estado": sesion.estado,
                    "tiempo_pausa_acumulado": sesion.tiempo_pausa_acumulado or 0,
                    "hora_pausa": sesion.hora_pausa.isoformat() if (sesion.estado == 'PAUSADO' and sesion.hora_pausa) else None
                }
            })

        return api_success(data={"session": None})
    except Exception as e:
        return api_error(str(e), status_code=500)


@pulido_bp.route('/api/pulido/admin/sesiones', methods=['GET'])
@require_role(ROLES_ADMIN_PULIDO)
def listar_sesiones_pulido_admin():
    """
    Panel de admin (plan 2026-08-31): todas las sesiones de Pulido activas/
    pausadas de CUALQUIER operaria, para supervisión en tablets compartidas.
    Pausar/reanudar desde el panel reutilizan /api/pulido/pausar y
    /api/pulido/reanudar (ya funcionan por id_pulido, sin candado de dueño);
    corregir un reporte reutiliza el POST /api/pulido normal -- el Ownership
    Guard ya deja pasar a roles admin/jefe preservando el responsable
    original (ver AuditService.resolver_y_validar_propietario).
    """
    try:
        return api_success(data={'sesiones': PulidoService.listar_sesiones_activas()})
    except Exception as e:
        logger.error(f"❌ Error listando sesiones de Pulido (admin): {e}")
        return api_error(str(e), status_code=500)


@pulido_bp.route('/api/pulido/tareas_pendientes', methods=['GET'])
def get_pulido_tareas_pendientes():
    try:
        responsable = request.args.get('responsable')
        if not responsable:
            return api_error("Falta responsable", status_code=400)

        tareas = ProduccionPulido.query.filter(
            ProduccionPulido.responsable == responsable,
            ProduccionPulido.estado.in_(['PENDIENTE', 'PAUSADO_COLA'])
        ).order_by(ProduccionPulido.id.desc()).all()

        return api_success(data={
            "tareas": [{
                "id_pulido": t.id_pulido,
                "codigo": t.codigo,
                "lote": t.lote,
                "orden_produccion": t.orden_produccion,
                "estado": t.estado
            } for t in tareas]
        })
    except Exception as e:
        return api_error(str(e), status_code=500)

@pulido_bp.route('/api/pulido/pausar', methods=['POST'])
@require_role(ROLES_PULIDO)
def pausar_pulido():
    """Registra el inicio de la pausa en el servidor."""
    data = request.json
    id_pulido = data.get('id_pulido')
    try:
        registro = ProduccionPulido.query.filter_by(id_pulido=id_pulido).first()
        if not registro: return api_error("No encontrado", status_code=404)

        # Blindaje: Forzar timestamp de Colombia (Bogotá)
        ahora = get_colombia_time()

        # Si el frontend envía una hora específica, intentar usarla para la parte de tiempo
        hora_front = data.get('hora_pausa')
        if hora_front and ':' in hora_front:
            try:
                h, m = hora_front.split(':')
                ahora = ahora.replace(hour=int(h), minute=int(m), second=0, microsecond=0)
            except: pass

        registro.estado = 'PAUSADO'
        registro.hora_pausa = ahora.replace(tzinfo=None) # Guardar como naive Bogota
        db.session.add(registro)
        db.session.commit()

        logger.debug(f" [PAUSA] Actividad {id_pulido} pausada a las {registro.hora_pausa}")
        return api_success()
    except Exception as e:
        db.session.rollback()
        return api_error(str(e), status_code=500)

@pulido_bp.route('/api/pulido/reanudar', methods=['POST'])
@require_role(ROLES_PULIDO)
def reanudar_pulido():
    """Calcula el tiempo de la pausa y lo suma al acumulador."""
    data = request.json
    id_pulido = data.get('id_pulido')
    try:
        registro = ProduccionPulido.query.filter_by(id_pulido=id_pulido).first()
        if not registro: return api_error("No encontrado", status_code=404)

        if registro.hora_pausa:
            ahora = get_colombia_time()

            # Si el frontend envía una hora específica de reanudación
            hora_front = data.get('hora_reanudar')
            if hora_front and ':' in hora_front:
                try:
                    h, m = hora_front.split(':')
                    ahora = ahora.replace(hour=int(h), minute=int(m), second=0, microsecond=0)
                except: pass

            ahora_naive = ahora.replace(tzinfo=None)
            diferencia = ahora_naive - registro.hora_pausa
            segundos_pausa = int(diferencia.total_seconds())
            if segundos_pausa < 0: segundos_pausa = 0 # Evitar pausas negativas por drift
            
            registro.tiempo_pausa_acumulado = (registro.tiempo_pausa_acumulado or 0) + segundos_pausa

        registro.estado = 'TRABAJANDO'
        registro.hora_pausa = None
        db.session.commit()
        return api_success(data={"acumulado": registro.tiempo_pausa_acumulado})
    except Exception as e:
        db.session.rollback()
        return api_error(str(e), status_code=500)

@pulido_bp.route('/api/pulido/swap_task', methods=['POST'])
@require_role(ROLES_PULIDO)
def swap_pulido_task():
    try:
        data = request.get_json()
        responsable = str(data.get('responsable') or "")
        id_raw = data.get('id_pulido')

        # Registro de depuración para Render
        logger.debug(f" [SWAP-DEBUG] Recibido id_pulido: {id_raw} (Tipo: {type(id_raw)})")

        # Blindaje definitivo contra [object Object] o diccionarios
        if isinstance(id_raw, dict):
            id_nuevo = str(id_raw.get('id_pulido') or "")
        elif str(id_raw) == "[object Object]":
            logger.error(" [SWAP-ERROR] El frontend envió un string '[object Object]'")
            return api_error("ID de tarea inválido (object)", status_code=400)
        else:
            id_nuevo = str(id_raw or "")

        if not responsable or not id_nuevo or id_nuevo == "None":
            return api_error("Datos incompletos o ID nulo", status_code=400)

        # 1. Pausar TODO lo que esté TRABAJANDO para este operario
        # Usamos la hora de Colombia para asegurar consistencia en el registro de pausa
        now = get_colombia_time()
        trabajos_activos = ProduccionPulido.query.filter(
            ProduccionPulido.responsable == responsable,
            ProduccionPulido.estado == 'TRABAJANDO'
        ).all()
        
        for t in trabajos_activos:
            t.estado = 'PAUSADO_COLA'
            t.hora_pausa = now
            db.session.add(t)
            
        # 2. Activar la nueva tarea
        nueva_tarea = ProduccionPulido.query.filter_by(id_pulido=id_nuevo).first()
        if nueva_tarea:
            nueva_tarea.estado = 'TRABAJANDO'
            nueva_tarea.hora_pausa = None # Limpiar pausa para reanudación
            db.session.add(nueva_tarea)
            
        db.session.commit()
        return api_success(message="Intercambio realizado con éxito")
    except Exception as e:
        db.session.rollback()
        logger.error(f"[Pulido-Swap] Error crítico: {e}")
        return api_error(str(e), status_code=500)
@pulido_bp.route('/api/pulido/historial', methods=['GET'])
@require_role(ROLES_PULIDO)
def get_pulido_historial():
    """
    Endpoint para el historial detallado de Pulido con filtros dinámicos.
    """
    try:
        # 1. Parámetros
        f_inicio = request.args.get('fecha_inicio', '')
        f_fin = request.args.get('fecha_fin', '')
        operario = request.args.get('operario', '')
        id_codigo = request.args.get('id_codigo', '')

        # 2. Base Query con Casts explícitos para evitar OID 25 (TEXT)
        sql = """
            SELECT 
                id,
                id_pulido::TEXT as id_pulido,
                fecha,
                codigo::TEXT as codigo,
                responsable::TEXT as responsable,
                cantidad_real,
                pnc_inyeccion,
                pnc_pulido,
                hora_inicio,
                hora_fin,
                orden_produccion::TEXT as orden_produccion,
                observaciones::TEXT as observaciones,
                cantidad_recibida
            FROM db_pulido
            WHERE 1=1
        """
        params = {}

        # 3. Filtros Dinámicos
        if f_inicio and f_fin:
            sql += " AND CAST(fecha AS DATE) BETWEEN :f_inicio AND :f_fin"
            params['f_inicio'] = f_inicio
            params['f_fin'] = f_fin
        elif f_inicio:
            sql += " AND CAST(fecha AS DATE) >= :f_inicio"
            params['f_inicio'] = f_inicio
        elif f_fin:
            sql += " AND CAST(fecha AS DATE) <= :f_fin"
            params['f_fin'] = f_fin

        if operario and operario.upper() != 'TODOS':
            sql += " AND UPPER(TRIM(responsable)) LIKE :operario"
            params['operario'] = f"%{operario.strip().upper()}%"

        if id_codigo and id_codigo.upper() != 'TODOS':
            sql += " AND UPPER(TRIM(codigo)) LIKE :codigo"
            params['codigo'] = f"%{id_codigo.strip().upper()}%"

        sql += " ORDER BY fecha DESC, id DESC"

        # 4. Ejecución y Conversión Inmediata a Diccionario (Clean Data)
        result = db.session.execute(text(sql), params)
        resultados = [dict(row._mapping) for row in result]

        # 5. Batch Pre-fetch de Revueltos (v5.2 - Zero queries in loop)
        all_ids = [str(r.get('id_pulido') or '').strip() for r in resultados]
        all_ids = [pid for pid in all_ids if pid]  # Filtrar vacíos

        revueltos_map = {}  # { id_pulido: [ {id_codigo, cantidad}, ... ] }
        if all_ids:
            # Una sola consulta para TODOS los revueltos
            placeholders = ', '.join([f':pid_{i}' for i in range(len(all_ids))])
            sql_revs = f"SELECT id_pulido::TEXT as id_pulido, id_codigo::TEXT as id_codigo, COALESCE(cantidad, 0) as cantidad FROM db_bujes_revueltos WHERE id_pulido IN ({placeholders})"
            params_revs = {f'pid_{i}': pid for i, pid in enumerate(all_ids)}
            revs_raw = db.session.execute(text(sql_revs), params_revs)
            for rv in revs_raw:
                rv_dict = dict(rv._mapping)
                pid = str(rv_dict['id_pulido'])
                if pid not in revueltos_map:
                    revueltos_map[pid] = []
                revueltos_map[pid].append(rv_dict)

        # 6. Mapeo para el Frontend (cero consultas en el bucle)
        data = []
        for r in resultados:
            p_id = str(r.get('id_pulido') or '').strip()

            # Lookup directo en memoria
            revs = revueltos_map.get(p_id, [])
            det_revueltos = ""
            if revs:
                det_revueltos = " | REVUELTOS: " + ", ".join([f"{str(rv['id_codigo'])}({float(rv['cantidad'] or 0)})" for rv in revs])

            obs = str(r.get('observaciones') or '').strip()
            detalle_final = f"Obs: {obs}{det_revueltos}" if obs else det_revueltos.strip(" | ")

            data.append({
                'id': int(r['id']),
                'Fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
                'Tipo': 'PULIDO',
                'Producto': str(r['codigo'] or ''),
                'Responsable': str(r['responsable'] or 'SISTEMA'),
                'cantidad_real': float(r['cantidad_real'] or 0),
                'Cant': float(r['cantidad_real'] or 0),
                'Orden': str(r['orden_produccion'] or p_id or '-'),
                'Extra': str(f"OP: {r['orden_produccion'] or ''}"),
                'Detalle': str(detalle_final.strip()),
                'HORA_INICIO': r['hora_inicio'].strftime('%H:%M') if r['hora_inicio'] else '',
                'HORA_FIN': r['hora_fin'].strftime('%H:%M') if r['hora_fin'] else '',
                'hoja': 'db_pulido',
                'fila': int(r['id']),
                'cantidad_recibida': float(r['cantidad_recibida'] or 0),
                'pnc_inyeccion': int(r['pnc_inyeccion'] or 0),
                'pnc_pulido': int(r['pnc_pulido'] or 0)
            })

        return api_success(data=data)

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error en api/pulido/historial: {e}")
        return api_error(str(e), status_code=500)

@pulido_bp.route('/api/pulido/stats', methods=['GET'])
@require_role(ROLES_PULIDO)
def get_pulido_stats():
    # Placeholder
    return api_success(message="Estadísticas de pulido (WIP)")


def _generar_excel_pulido_task(task_id, f_inicio, f_fin, operario, id_codigo):
    """
    Trabajo de fondo de exportar_excel_pulido: arma el Excel completo (consulta
    + estilos openpyxl) fuera del hilo HTTP. Corre dentro del app_context que
    le da task_runner.run_in_background.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        # 2. Query con casts
        sql = """
            SELECT
                id, id_pulido::TEXT as id_pulido, fecha,
                codigo::TEXT as codigo, responsable::TEXT as responsable,
                cantidad_real, pnc_inyeccion, pnc_pulido,
                hora_inicio, hora_fin, tiempo_total_minutos,
                orden_produccion::TEXT as orden_produccion,
                observaciones::TEXT as observaciones,
                cantidad_recibida
            FROM db_pulido
            WHERE 1=1
        """
        params = {}

        if f_inicio and f_fin:
            sql += " AND CAST(fecha AS DATE) BETWEEN :f_inicio AND :f_fin"
            params['f_inicio'] = f_inicio
            params['f_fin'] = f_fin
        elif f_inicio:
            sql += " AND CAST(fecha AS DATE) >= :f_inicio"
            params['f_inicio'] = f_inicio
        elif f_fin:
            sql += " AND CAST(fecha AS DATE) <= :f_fin"
            params['f_fin'] = f_fin

        if operario and operario.upper() != 'TODOS':
            sql += " AND UPPER(TRIM(responsable)) LIKE :operario"
            params['operario'] = f"%{operario.strip().upper()}%"

        if id_codigo and id_codigo.upper() != 'TODOS':
            sql += " AND UPPER(TRIM(codigo)) LIKE :codigo"
            params['codigo'] = f"%{id_codigo.strip().upper()}%"

        sql += " ORDER BY fecha DESC, id DESC"

        result = db.session.execute(text(sql), params)
        resultados = [dict(row._mapping) for row in result]

        # 3. Crear Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Pulido"

        # --- ESTILOS ---
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D5D8DC'),
            right=Side(style='thin', color='D5D8DC'),
            top=Side(style='thin', color='D5D8DC'),
            bottom=Side(style='thin', color='D5D8DC')
        )
        zebra_fill = PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid')
        pnc_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
        # Horario fuera de jornada (turno único 07:00-17:00, ver
        # PulidoService.validar_duracion_turno) -- casi siempre delatan un
        # AM/PM mal marcado en el picker de "MODO SATÉLITE" (input type=time,
        # que en celulares con formato 12h puede precargar AM/PM según la
        # hora del día en que la operaria abre el campo, no según el turno
        # real que está reportando).
        horario_sospechoso_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        data_align = Alignment(horizontal='center', vertical='center')
        text_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # --- CABECERA ---
        columnas = [
            'Fecha', 'Hora Inicio', 'Hora Fin', 'Tiempo Real (min)', 'Operario', 'Referencia',
            'Orden/OP', 'Cant. Recibida', 'Cantidad OK', 'PNC Iny', 'PNC Pul', 'Observaciones'
        ]
        for col_idx, titulo in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Ventana de jornada única de Pulido (07:00-17:00, sin turno
        # nocturno -- ver PulidoService.validar_duracion_turno). Se deja un
        # margen razonable para entradas tempranas/salidas con sobretiempo
        # antes de marcar la fila como sospechosa.
        from datetime import time as _time
        VENTANA_INICIO_MIN = _time(6, 0)
        VENTANA_FIN_MAX = _time(18, 0)

        # --- FILAS DE DATOS ---
        for row_idx, r in enumerate(resultados, 2):
            fecha_str = r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else ''
            h_inicio = r['hora_inicio'].strftime('%H:%M') if r['hora_inicio'] else ''
            h_fin = r['hora_fin'].strftime('%H:%M') if r['hora_fin'] else ''
            tiempo_real_min = round(float(r['tiempo_total_minutos'] or 0), 1)
            cant_recibida = float(r['cantidad_recibida'] or 0)
            cant_ok = int(r['cantidad_real'] or 0)
            pnc_i = int(r['pnc_inyeccion'] or 0)
            pnc_p = int(r['pnc_pulido'] or 0)

            hora_sospechosa = bool(
                (r['hora_inicio'] and r['hora_inicio'].time() < VENTANA_INICIO_MIN)
                or (r['hora_fin'] and r['hora_fin'].time() > VENTANA_FIN_MAX)
            )

            fila = [
                fecha_str,
                h_inicio,
                h_fin,
                tiempo_real_min,
                str(r['responsable'] or 'SISTEMA'),
                str(r['codigo'] or ''),
                str(r['orden_produccion'] or r['id_pulido'] or '-'),
                round(cant_recibida, 1) if cant_recibida != int(cant_recibida) else int(cant_recibida),
                cant_ok,
                pnc_i,
                pnc_p,
                str(r['observaciones'] or '')
            ]

            for col_idx, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = thin_border

                # Alineación: texto a la izquierda, números al centro
                if col_idx in (5, 6, 7, 12):  # Operario, Referencia, Orden, Observaciones
                    cell.alignment = text_align
                else:
                    cell.alignment = data_align

            # Cebreado (filas pares)
            if row_idx % 2 == 0:
                for col_idx in range(1, len(columnas) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = zebra_fill

            # Fuera de jornada (probable AM/PM mal marcado en MODO SATÉLITE):
            # resalta Hora Inicio y Hora Fin para que salte a la vista sin
            # tener que revisar turno por turno a mano.
            if hora_sospechosa:
                for col_idx in (2, 3):
                    ws.cell(row=row_idx, column=col_idx).fill = horario_sospechoso_fill
                    ws.cell(row=row_idx, column=col_idx).font = Font(bold=True, color='7A5B00')

            # Alerta PNC (fondo rojo claro si > 0)
            if pnc_i > 0:
                ws.cell(row=row_idx, column=10).fill = pnc_fill
                ws.cell(row=row_idx, column=10).font = Font(bold=True, color='C0392B')
            if pnc_p > 0:
                ws.cell(row=row_idx, column=11).fill = pnc_fill
                ws.cell(row=row_idx, column=11).font = Font(bold=True, color='C0392B')

        # --- AUTO-AJUSTE DE ANCHOS ---
        anchos = [12, 11, 11, 15, 22, 18, 18, 14, 13, 9, 9, 40]
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Inmovilizar primera fila
        ws.freeze_panes = 'A2'

        # 4. Generar archivo en memoria y volcarlo a disco: el BytesIO no
        # sobrevive a este hilo -- la descarga es un request HTTP aparte.
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fecha_archivo = datetime.now().strftime('%Y-%m-%d')
        filename = f"Historial_Pulido_{fecha_archivo}.xlsx"

        fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', prefix='pulido_')
        with os.fdopen(fd, 'wb') as f:
            f.write(output.getvalue())

        logger.debug(f"📊 [Excel] Exportando {len(resultados)} registros de Pulido")

        task_runner.set_completed(
            task_id, file_path=tmp_path, filename=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error exportando Excel Pulido (task {task_id}): {e}")
        import traceback
        logger.error(traceback.format_exc())
        task_runner.set_failed(task_id, str(e))


@pulido_bp.route('/api/pulido/exportar_excel', methods=['GET'])
@require_role(ROLES_PULIDO)
def exportar_excel_pulido():
    """
    Controller delgado: valida parámetros y delega la generación completa del
    Excel a un hilo de fondo (ver _generar_excel_pulido_task).
    """
    f_inicio = request.args.get('fecha_inicio', '')
    f_fin = request.args.get('fecha_fin', '')
    operario = request.args.get('operario', '')
    id_codigo = request.args.get('id_codigo', '')

    task_id = task_runner.create_task()
    app_obj = current_app._get_current_object()
    task_runner.run_in_background(
        task_id, app_obj, _generar_excel_pulido_task,
        f_inicio, f_fin, operario, id_codigo
    )

    return api_success(data={"task_id": task_id}, status_code=202)

# =============================================================
# NUEVO: Endpoint GET — Lotes Activos para Modo Lotes en Vivo
# Devuelve todos los lotes en estado ABIERTO_PRODUCCION.
# Sin locks: cualquier número de operarias puede consultarlo
# simultáneamente sin colisiones (solo lectura).
# =============================================================
@pulido_bp.route('/api/pulido/lotes_activos', methods=['GET'])
@require_role(ROLES_PULIDO)
def get_lotes_activos():
    """
    Retorna los lotes de db_trazabilidad_lotes con estado ABIERTO_PRODUCCION.
    Pulido lo consulta al entrar al Modo Lotes en Vivo para mostrar la lista táctil.
    """
    try:
        # from sqlalchemy import or_
        # lotes = db.session.query(TrazabilidadLote).filter(
        #     or_(
        #         # Flujo normal
        #         db.and_(
        #             TrazabilidadLote.estado_actual.in_(['ABIERTO_PRODUCCION', 'ABIERTO_PULIDO']),
        #             TrazabilidadLote.por_pulir > 0
        #         ),
        #         # Forzar visibilidad para lotes de prueba
        #         TrazabilidadLote.id_lote.ilike('%PRUEBA%'),
        #         TrazabilidadLote.id_lote.ilike('%9999%'),
        #         TrazabilidadLote.orden_produccion.ilike('%PRUEBA%'),
        #         TrazabilidadLote.orden_produccion.ilike('%9999%')
        #     )
        # ).order_by(TrazabilidadLote.fecha_creacion.desc()).all()

        # Flujo estricto de producción
        lotes = db.session.query(TrazabilidadLote).filter(
            TrazabilidadLote.estado_actual == 'ABIERTO_PRODUCCION',
            TrazabilidadLote.por_pulir > 0
        ).order_by(TrazabilidadLote.fecha_creacion.desc()).all()

        resultado = [{
            'id_lote'          : l.id_lote,
            'orden_produccion' : l.orden_produccion or 'SIN OP',
            'id_codigo'        : l.id_codigo,
            'maquina'          : l.maquina or '',
            'responsable'      : l.responsable or '',
            'fecha_creacion'   : l.fecha_creacion.strftime('%d/%m/%Y %H:%M') if l.fecha_creacion else '',
            'estado_actual'    : l.estado_actual,
            'cantidad_inyectada': l.cantidad_inyectada or 0,
            'por_pulir'        : l.por_pulir or 0
        } for l in lotes]

        return api_success(data={'lotes': resultado})

    except Exception as e:
        logger.error(f'❌ [Lotes Activos] Error: {e}')
        return api_error(str(e), status_code=500)


@pulido_bp.route('/api/pulido/saldo_por_op', methods=['GET'])
@require_role(ROLES_PULIDO)
def get_saldo_por_op():
    """Fase 7 (plan OP->WO): saldo real de 'por pulir' por OP+referencia --
    inyectado - pulido, solo OP con trazabilidad real (excluye 'SIN OP' y
    todo lo anterior al corte). Lógica en PulidoService.get_saldo_por_op."""
    try:
        return api_success(data={'saldo': PulidoService.get_saldo_por_op()})
    except Exception as e:
        logger.error(f'❌ [Saldo por OP] Error: {e}')
        return api_error(str(e), status_code=500)


@pulido_bp.route('/api/pulido/liquidar_lote', methods=['POST'])
@require_role(ROLES_PULIDO)
def liquidar_lote():
    try:
        data = request.get_json()
        if not data:
            return api_error("No data provided", status_code=400)

        id_lote = data.get('id_lote')
        from backend.utils.formatters import resolver_operario
        responsable = resolver_operario(data.get('responsable')) or 'Supervisor'

        if not id_lote:
            return api_error("Falta id_lote", status_code=400)

        # Buscamos todos los lotes de trazabilidad que coincidan con id_lote
        # Puede ser un ID de lote específico (con sufijo) o el prefijo base del lote
        lotes_traz = TrazabilidadLote.query.filter(
            (TrazabilidadLote.id_lote == id_lote) |
            (TrazabilidadLote.id_lote.like(f"{id_lote}-%"))
        ).all()

        if not lotes_traz:
            return api_error(f"Lote {id_lote} no encontrado", status_code=404)

        lotes_procesados = 0

        for lote_traz in lotes_traz:
            # Poner por_pulir a 0 y cambiar estado
            lote_traz.por_pulir = 0
            lote_traz.estado_actual = 'PENDIENTE_VALIDACION'
            lotes_procesados += 1

        db.session.commit()
        logger.info(f"⚡ [Liquidar Lote] Se liquidaron/cerraron {lotes_procesados} referencia(s) de {id_lote} por {responsable} (por_pulir establecido en 0 y estado -> PENDIENTE_VALIDACION).")
        return api_success(message=f"Lote {id_lote} cerrado administrativamente correctamente.")

    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error liquidando lote: {e}")
        return api_error(str(e), status_code=500)


@pulido_bp.route('/reporte_masivo', methods=['POST'])
@pulido_bp.route('/api/pulido/reporte_masivo', methods=['POST'])
@require_role(ROLES_PULIDO)
def reporte_masivo():
    try:
        data = request.get_json()
        if not data:
            return api_error("No data provided", status_code=400)

        hora_inicio_str = data.get('hora_inicio')
        hora_fin_str = data.get('hora_fin')
        from backend.utils.formatters import resolver_operario
        responsable = resolver_operario(data.get('responsable'))
        items = data.get('items', [])

        if not responsable or responsable == 'SISTEMA':
            return api_error("Falta el responsable", status_code=400)

        if not items:
            return api_error("No hay items para registrar", status_code=400)

        ahora = get_colombia_time()
        fecha_actual = ahora.date()

        for item in items:
            referencia_raw = item.get('referencia')
            if not referencia_raw:
                raise ValueError("Se requiere la referencia en todos los registros del lote")

            buenos = float(item.get('buenos') or 0)
            rev_total = sum(float(r.get('cantidad', 0)) for r in item.get('revueltos', []))
            
            # Filtro de Seguridad Backend (Anti-Basura)
            if (buenos + rev_total) <= 0:
                logger.debug(f"Omitiendo registro de {referencia_raw} porque la cantidad total es cero.")
                continue

            # ── Blindaje: se guarda la referencia tal cual la dictó la operaria,
            #    con su prefijo original o sin ninguno; no se infiere división ──
            referencia = preservar_o_normalizar_prefijo(referencia_raw)
            # referencia_sin_prefijo sólo se usa para los cruces con db_distribucion_op_pedidos
            referencia_sin_prefijo = normalizar_codigo(referencia_raw)
            op = item.get('op') or 'SIN OP'
            lote = item.get('lote') or 'SIN LOTE'

            id_pulido = f"PUL-VOZ-{uuid.uuid4().hex[:8].upper()}"
            registro = ProduccionPulido(
                id_pulido=id_pulido,
                fecha=fecha_actual,
                fecha_registro=ahora,
                codigo=referencia,
                responsable=responsable,
                cantidad_real=int(buenos),
                pnc_pulido=0,
                pnc_inyeccion=0,
                orden_produccion=op,
                lote=lote,
                estado='FINALIZADO',
                departamento='PULIDO',
                cantidad_recibida=buenos,
                almacen_destino='P. TERMINADO',
                observaciones='Reporte Masivo por Voz Fin de Turno'
            )

            # Tiempos e indicadores de tiempo se ignoran para el reporte masivo/de lotes
            registro.hora_inicio = None
            registro.hora_fin = None
            registro.duracion_segundos = 0
            registro.tiempo_total_minutos = 0.0
            registro.segundos_por_unidad = 0.0

            db.session.add(registro)

            # --- MES PASO 2: db_productos NO se toca aquí. ---
            # La mutación de inventario ocurre EXCLUSIVAMENTE en el módulo de Validación (Paso 3).
            # Esta función solo persiste el reporte físico (db_pulido) y propaga al lote de trazabilidad.

            # --- MES PASO 2: Propagar al Lote de Trazabilidad (si viene en Modo Lotes en Vivo) ---
            id_lote_ref = item.get('id_lote')
            if id_lote_ref:
                lote_traz = db.session.get(TrazabilidadLote, id_lote_ref)
                if lote_traz:
                    # ---------------------------------------------------------
                    # Descuento en Lote Continuo (TrazabilidadLote.por_pulir)
                    # ---------------------------------------------------------
                    revueltos_items = item.get('revueltos', [])
                    rev_total = sum(float(r.get('cantidad', 0)) for r in revueltos_items)
                    procesado_lote = buenos + rev_total
                    
                    lote_traz.por_pulir = max(0, (lote_traz.por_pulir or 0) - int(procesado_lote))
                    
                    if lote_traz.por_pulir <= 0:
                        lote_traz.estado_actual = 'PENDIENTE_VALIDACION'
                        logger.info(f'🟡 [Trazabilidad] Lote {id_lote_ref} completó por_pulir -> PENDIENTE_VALIDACION por operaria {responsable}')
                    else:
                        logger.debug(f'🔄 [Trazabilidad] Lote {id_lote_ref} registrado. Quedan {lote_traz.por_pulir} por pulir.')
                else:
                    logger.warning(f'⚠️ [Trazabilidad] id_lote {id_lote_ref} no encontrado en db_trazabilidad_lotes')

            # 4. Registrar Revueltos Masivos
            revueltos_items = item.get('revueltos', [])
            if revueltos_items and id_lote_ref:
                for r_item in revueltos_items:
                    r_cod_con_fr = preservar_o_normalizar_prefijo(r_item.get('id_codigo'))
                    r_cant = float(r_item.get('cantidad') or 0)
                    if r_cod_con_fr and r_cant > 0:
                        db.session.add(BujeRevuelto(
                            id_bujes_revueltos=uuid.uuid4().hex[:8],
                            id_pulido=id_pulido,
                            id_lote=id_lote_ref,
                            id_codigo=r_cod_con_fr,
                            cantidad=int(r_cant),
                            responsable=responsable
                        ))

            # --- Propagación FIFO ---
            if op and str(op).strip() != 'SIN OP':
                from backend.models.sql_models import DistribucionOpPedidos
                # ── Blindaje Dual: cruce FIFO con db_distribucion_op_pedidos usa SIN prefijo ──
                op_limpia = str(op).strip()
                codigo_limpio = referencia_sin_prefijo  # normalizado arriba, sin FR-
                
                cubetas = db.session.query(DistribucionOpPedidos).filter(
                    DistribucionOpPedidos.op_world_office == op_limpia,
                    DistribucionOpPedidos.codigo_producto == codigo_limpio
                ).order_by(DistribucionOpPedidos.id_distribucion.asc()).all()

                piezas_por_repartir = float(buenos)

                if not cubetas and piezas_por_repartir > 0:
                    pedido_asoc = db.session.query(DistribucionOpPedidos.id_pedido).filter(
                        DistribucionOpPedidos.op_world_office == op_limpia
                    ).first()
                    id_pedido_final = pedido_asoc[0] if (pedido_asoc and pedido_asoc[0]) else f"PED-IMPREVISTO-{op_limpia}"
                    
                    logger.info(f"[PULIDO-MASIVO-CONTINGENCIA] Creando cubeta temporal para OP: {op_limpia}, Producto: {codigo_limpio}")
                    nueva_cubeta = DistribucionOpPedidos(
                        op_world_office=op_limpia,
                        id_pedido=id_pedido_final,
                        codigo_producto=codigo_limpio,
                        cant_requerida=piezas_por_repartir,
                        cant_inyectada=piezas_por_repartir,
                        cant_pulida=piezas_por_repartir,
                        cant_ensamblada=0,
                        cant_alistada=0
                    )
                    db.session.add(nueva_cubeta)
                    db.session.flush()
                    cubetas = [nueva_cubeta]
                    piezas_por_repartir = 0.0

                for cubeta in cubetas:
                    if piezas_por_repartir <= 0:
                        break
                    falta = max(0, (cubeta.cant_requerida or 0) - (cubeta.cant_pulida or 0))
                    if falta > 0:
                        if piezas_por_repartir >= falta:
                            cubeta.cant_pulida = (cubeta.cant_pulida or 0) + falta
                            piezas_por_repartir -= falta
                        else:
                            cubeta.cant_pulida = (cubeta.cant_pulida or 0) + piezas_por_repartir
                            piezas_por_repartir = 0

        db.session.commit()
        return api_success(message=f"Se registraron con éxito {len(items)} reportes del lote.")

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error en reporte_masivo: {e}")
        return api_error(str(e), status_code=500)


@pulido_bp.route('/api/pnc/registrar_pulido', methods=['POST'])
@require_role(ROLES_PULIDO)
def registrar_pnc_pulido():
    """
    Registra (o limpia) el desglose de PNC de Pulido en db_pnc_pulido
    y sincroniza pnc_pulido + cantidad_real en db_pulido.
    Body JSON:
      id_pulido - ID de la sesión de pulido
      id_codigo - Código del producto
      defectos  - {
          "Porosidad / Burbujas": N,
          "Marcas de Lijado / Rayones": N,
          "Desgaste Excesivo / Deformación": N,
          "Brillo Insuficiente": N
        }
    """
    try:
        data = request.get_json() or {}
        id_pulido = data.get('id_pulido')
        id_codigo = data.get('id_codigo')
        defectos  = data.get('defectos', {})

        if not id_pulido or not id_codigo:
            return api_error("id_pulido e id_codigo son obligatorios", status_code=400)

        id_cod = normalizar_codigo(id_codigo)

        # Eliminar registro previo para este turno + producto
        db.session.query(PncPulido).filter_by(id_pulido=id_pulido, codigo=id_cod).delete()

        # Mapear los 4 criterios específicos de Pulido
        porosidad = float(defectos.get("Porosidad / Burbujas", 0) or 0)
        rayones   = float(defectos.get("Marcas de Lijado / Rayones", 0) or 0)
        desgaste  = float(defectos.get("Desgaste Excesivo / Deformación", 0) or 0)
        brillo    = float(defectos.get("Brillo Insuficiente", 0) or 0)
        total_pnc = porosidad + rayones + desgaste + brillo

        prod_pul = db.session.query(ProduccionPulido).filter_by(id_pulido=id_pulido).first()

        if total_pnc > 0:
            # La merma se atribuye a la operaria del turno de pulido. Si el
            # turno no existe o no tiene responsable identificable, se rechaza:
            # persistir la merma sin dueño reabre el vacío de trazabilidad.
            try:
                operaria_responsable = PulidoService.resolver_operaria_responsable(prod_pul)
            except ValueError as ve:
                return api_error(str(ve), status_code=400, code="RESPONSABLE_REQUERIDO")

            criterio_str = (
                f"Porosidad/Burbujas: {int(porosidad)}, "
                f"Rayones: {int(rayones)}, "
                f"Desgaste/Deformación: {int(desgaste)}, "
                f"Brillo Insuficiente: {int(brillo)}"
            )
            nuevo_pnc = PncPulido(
                id_pnc_pulido=uuid.uuid4().hex[:8],
                id_pulido=id_pulido,
                codigo=id_cod,
                cantidad=total_pnc,
                criterio=criterio_str,
                responsable=operaria_responsable
            )
            db.session.add(nuevo_pnc)

            if prod_pul:
                prod_pul.pnc_pulido = int(round(total_pnc))
                prod_pul.criterio_pnc_pulido = criterio_str
                cant_bruta = float(prod_pul.cantidad_recibida or prod_pul.cantidad_real or 0)
                prod_pul.cantidad_real = max(0, int(round(cant_bruta - total_pnc)))

            db.session.commit()
            logger.info(f"✅ PNC Pulido registrado para {id_cod} en {id_pulido}: Total={total_pnc}")
            return api_success(data={"total_pnc": total_pnc}, message="PNC de Pulido registrado en db_pnc_pulido")
        else:
            if prod_pul:
                prod_pul.pnc_pulido = 0
                prod_pul.criterio_pnc_pulido = ""
            db.session.commit()
            return api_success(data={"total_pnc": 0}, message="Sin defectos de PNC para Pulido")

    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error registrando PNC Pulido: {e}")
        return api_error(str(e), status_code=500)

# Eliminado endpoint duplicado /api/pulido/liquidar_lote

# Eliminado 2026-08-04: flujo MES /api/mes/pulido/* (estado/iniciar/finalizar/
# pausar/reanudar/resumen_pausas) — código huérfano, sin ningún caller en
# frontend/ (verificado por grep). Reemplazado por el flujo real que usa
# pulido.js: /api/pulido (POST), /api/pulido/session_active, /api/pulido/pausar,
# /api/pulido/reanudar, /api/pulido/swap_task.


@pulido_bp.route('/api/pulido/ultimo_registro/<responsable>', methods=['GET'])
def get_ultimo_registro_pulido_legacy(responsable):
    """
    Variante legacy (path param en vez de query string) de
    /api/pulido/ultimo_registro. Movida desde app.py con el bug corregido:
    la rama en que SÍ se encuentra un registro no retornaba nada, lo que
    causaba un 500 de Flask ('view function did not return a valid
    response'). Se replica el shape de respuesta de la variante hermana
    (get_ultimo_registro_pulido, arriba) para que ambas sean consistentes.
    """
    try:
        registro = (
            ProduccionPulido.query
            .filter(ProduccionPulido.responsable == responsable)
            .order_by(ProduccionPulido.id.desc())
            .first()
        )

        if not registro:
            return api_success(data={'registro': None})

        hora_ref = registro.hora_fin or registro.hora_inicio
        if registro.fecha and hora_ref:
            try:
                fecha_hora_str = f"{registro.fecha.strftime('%d/%m/%Y')} {hora_ref.strftime('%H:%M')}"
            except Exception:
                fecha_hora_str = str(registro.fecha)
        else:
            fecha_hora_str = registro.fecha.strftime('%d/%m/%Y') if registro.fecha else '—'

        return api_success(data={
            'registro': {
                'fecha_hora': fecha_hora_str,
                'codigo_producto': registro.codigo or '—',
                'cantidad': float(registro.cantidad_real or 0),
                'cantidad_aprobada': float(registro.cantidad_real or 0),
                'piezas': float(registro.cantidad_real or 0)
            }
        })

    except Exception as e:
        logger.error(f"Error en get_ultimo_registro_pulido_legacy: {e}")
        return api_error(str(e), status_code=500)
