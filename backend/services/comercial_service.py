import logging
import re
from datetime import datetime, date
from sqlalchemy import text
from backend.core.sql_database import db

logger = logging.getLogger(__name__)

# Aislamiento estricto por igualdad exacta (no LIKE): permite uso de indice B-Tree
# sobre db_ventas.vendedor y evita falsos positivos entre nombres similares.
FILTRO_SCOPE_ROL = "(:es_comercial = FALSE OR UPPER(TRIM(COALESCE(v.vendedor, ''))) = UPPER(TRIM(:vendedor_user)))"
FILTRO_VENDEDOR_OPCIONAL = "(:vendedor_filtro = '' OR UPPER(TRIM(COALESCE(v.vendedor, ''))) = UPPER(TRIM(:vendedor_filtro)))"
# "Ventas Totales" no debe mezclar Pedidos (db_ventas.clasificacion='pedido'): sin este filtro,
# toda cifra de este servicio suma venta+pedido en un solo numero (confirmado con datos reales:
# Andres/2025 daba $5,309,314,039 = venta $2,590,839,285 + pedido $2,718,474,754).
FILTRO_SOLO_VENTA = "UPPER(TRIM(COALESCE(v.clasificacion, ''))) = 'VENTA'"

# Mapeo estático de configuración: agrupa ciudades (tal como llegan en db_ventas.zona,
# poblado desde Ciudad_Encabezado de World Office) en macro-zonas comerciales.
# Construido a partir de `SELECT UPPER(TRIM(zona)), COUNT(*) FROM db_ventas GROUP BY 1`
# corrido contra la base real (2026-07-29), cubriendo el listado completo de ciudades
# con ventas registradas. Cualquier ciudad nueva que no esté aquí cae en 'SIN ZONA'.
# Ambigüedades marcadas con comentario: ubicación asignada por mejor criterio geográfico,
# no verificada 1:1 contra el maestro de clientes — validar si la precisión de zona importa.
MAPEO_ZONAS = {
    'Llanos': [
        'VILLAVICENCIO', 'YOPAL', 'GRANADA', 'ACACIAS', 'ACACÍAS', 'PUERTO LOPEZ', 'PUERTO LÓPEZ',
        'PUERTO GAITAN', 'PUERTO GAITÁN', 'SAN JOSE DEL GUAVIARE', 'SAN JOSÉ DEL GUAVIARE',
        'TAME', 'SARAVENA', 'AGUAZUL',
        'VILLANUEVA',  # Villanueva, Casanare (sin espacio en el dato real)
    ],
    'Costa': [
        'BARRANQUILLA', 'CARTAGENA', 'SANTA MARTA', 'MONTERIA', 'MONTERÍA', 'SINCELEJO',
        'VALLEDUPAR', 'RIOHACHA', 'SOLEDAD', 'AGUACHICA', 'AGUSTIN CODAZZI', 'AGUSTÍN CODAZZI',
        'CIENEGA', 'CIÉNEGA',
        'VILLA NUEVA',  # ambiguo: se asume Villanueva, La Guajira (con espacio en el dato real)
    ],
    'Eje Cafetero': [
        'MANIZALES', 'PEREIRA', 'ARMENIA', 'DOSQUEBRADAS', 'LA VIRGINIA', 'RISARALDA',
        'QUINCHIA', 'QUINCHÍA',
    ],
    'Antioquia': [
        'MEDELLIN', 'MEDELLÍN', 'BELLO', 'ITAGUI', 'ENVIGADO', 'RIONEGRO', 'SABANETA',
        'LA ESTRELLA', 'MARINILLA',
    ],
    'Centro': [
        'BOGOTA', 'BOGOTA D.C.', 'SOACHA', 'CHIA', 'CHÍA', 'ZIPAQUIRA', 'ZIPAQUIRÁ',
        'FACATATIVA', 'MOSQUERA', 'FUNZA', 'LA CALERA', 'VILLA DE SAN DIEGO DE UBATE',
        'CAJICA', 'CAJICÁ', 'VILLETA', 'LA MESA',
        'MADRID',  # asumido Madrid, Cundinamarca (sin sufijo de país, a diferencia de las capitales internacionales)
        'TUNJA', 'CHIQUINQUIRA', 'CHIQUINQUIRÁ', 'SAMACA', 'SAMACÁ', 'RAMIRIQUI', 'RAMIRIQUÍ',
        'MIRAFLORES', 'SOGAMOSO', 'NOBSA', 'CHOACHI', 'CHOACHÍ', 'DUITAMA',
        'FUSAGASUGA', 'FUSAGASUGÁ',
        'SAN JUAN',  # ambiguo por nombre genérico, bajo volumen
    ],
    'Santanderes': [
        'BUCARAMANGA', 'CUCUTA', 'CÚCUTA', 'FLORIDABLANCA', 'PIEDECUESTA', 'GIRON', 'GIRÓN',
        'BARRANCABERMEJA', 'SAN ALBERTO',  # San Alberto, Cesar: corredor Magdalena Medio junto a Barrancabermeja
        'BARBOSA',  # ambiguo: homónimo con Barbosa, Antioquia; se asume Barbosa, Santander
    ],
    'Suroccidente': [
        'CALI', 'PALMIRA', 'POPAYAN', 'POPAYÁN', 'PASTO', 'TULUA', 'TULUÁ', 'BUENAVENTURA',
        'BUGA', 'GUADALAJARA DE BUGA',  # nombre oficial completo de Buga tal como llega en el dato real
        'YUMBO', 'JAMUNDI', 'JAMUNDÍ', 'SANTANDER DE QUILICHAO', 'IPIALES', 'SAMANIEGO',
        'PIENDAMO', 'PIENDAMÓ',
        'NARIÑO', 'NARINO',  # departamento usado como zona genérica en algunos registros
    ],
    'Tolima Huila': [
        'IBAGUE', 'IBAGUÉ', 'NEIVA', 'ESPINAL', 'GIRARDOT', 'PITALITO', 'PLANADAS',
        'ANZOATEGUI',  # Anzoátegui, Tolima
        'LA DORADA', 'PUERTO TRIUNFO',  # corredor Magdalena Medio (La Dorada-Caldas / Puerto Triunfo-Antioquia)
        'FLORENCIA', 'CARTAGENA DEL CHAIRA', 'CARTAGENA DEL CHAIRÁ',  # Caquetá; NO confundir con Cartagena (Costa)
    ],
    'Internacional': [
        'QUITO-ECUADOR', 'LIMA-PERU', 'LIMA-PERÚ', 'SANTO DOMINGO-', 'SAN PEDRO SULA',
        'LA PAZ BOLIVIA', 'PANAMA', 'PANAMÁ', 'SANTIAGO DE CHILE',
    ],
}
# Inverso precalculado: ciudad -> zona, para etiquetar cada cliente por región una sola vez.
CIUDAD_A_ZONA = {ciudad: zona for zona, ciudades in MAPEO_ZONAS.items() for ciudad in ciudades}


class ComercialHistoricoService:
    """
    Servicio de extracción y analítica comercial histórica.
    Lee vendedor y zona de forma plana desde las columnas nativas de db_ventas
    (pobladas por World Office: Nombres_tercero_interno / Ciudad_Encabezado).
    Garantiza aislamiento estricto por rol y fidelidad contable restando
    devoluciones (Notas Crédito) directamente en PostgreSQL.
    """

    @staticmethod
    def _expr_ajustado_nc(columna: str) -> str:
        """CASE que invierte el signo de `columna` cuando el registro es una Nota Crédito."""
        return f"""
            CASE
                WHEN UPPER(TRIM(COALESCE(v.clasificacion, ''))) LIKE '%NC%'
                  OR UPPER(TRIM(COALESCE(v.documento, ''))) LIKE '%NC%'
                THEN -ABS(COALESCE(v.{columna}, 0))
                ELSE COALESCE(v.{columna}, 0)
            END
        """

    @staticmethod
    def _resolver_alias_vendedor(username: str) -> str:
        """
        Resuelve el valor a comparar contra db_ventas.vendedor (match exacto):
        prioriza db_usuarios.alias_vendedor_wo (nombre tal como lo escribe World
        Office) y cae al nombre de login si el alias no está configurado.

        Se busca por `username` (no por user_id de sesión): ningún endpoint de
        login en auth_routes.py escribe session['user_id']/['usuario_id'], por lo
        que ese id siempre llega en 0 y resolvería, para cualquier usuario, el
        alias del registro id=0 en vez del propio.
        """
        try:
            from backend.models.sql_models import Usuario
            if username:
                usuario = Usuario.query.filter_by(username=str(username)).first()
                if usuario and usuario.alias_vendedor_wo and usuario.alias_vendedor_wo.strip():
                    return usuario.alias_vendedor_wo.strip()
        except Exception as e:
            logger.warning(f"[COMERCIAL_SERVICE] No se pudo resolver alias_vendedor_wo para username={username!r}: {e}")
        return str(username or '')

    @staticmethod
    def _resolver_fecha_corte(fecha_corte_raw) -> date:
        """
        Valida y convierte fecha_corte a un date real. Si no viene o es invalida,
        usa hoy. Nunca deja pasar un string sin validar hacia la consulta SQL.
        """
        if fecha_corte_raw:
            try:
                return datetime.strptime(str(fecha_corte_raw)[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                logger.warning(f"[COMERCIAL_SERVICE] fecha_corte inválida recibida: {fecha_corte_raw!r}. Usando hoy.")
        return datetime.now().date()

    @staticmethod
    def obtener_analitica_historica(user_id: int, username: str, user_role: str, start_year: int = 2024, end_year: int = 2026) -> dict:
        """
        Extrae datos consolidados de ventas (Año, Mes, Zona, Cliente) entre start_year y end_year.
        Resta devoluciones/NC. Vendedor y zona se leen directo de db_ventas (sin JOIN).
        """
        role_upper = str(user_role or '').strip().upper()
        es_global = role_upper in ['ADMIN', 'ADMINISTRACION', 'ADMINISTRADOR', 'GERENCIA']
        es_comercial = not es_global

        ventas_expr = ComercialHistoricoService._expr_ajustado_nc('total_ingresos')
        cantidad_expr = ComercialHistoricoService._expr_ajustado_nc('cantidad')
        vendedor_scope = ComercialHistoricoService._resolver_alias_vendedor(username)

        params = {
            'start_year': int(start_year),
            'end_year': int(end_year),
            'es_comercial': es_comercial,
            'vendedor_user': vendedor_scope,
            'vendedor_filtro': ''  # sin acotar adicionalmente en esta vista
        }

        query_resumen = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                COUNT(v.id)::INTEGER AS total_transacciones
            FROM db_ventas v
            WHERE v.fecha >= MAKE_DATE(:start_year, 1, 1)
              AND CAST(v.fecha AS DATE) <= MAKE_DATE(:end_year, 12, 31)
              AND {FILTRO_SCOPE_ROL}
              AND {FILTRO_VENDEDOR_OPCIONAL}
              AND {FILTRO_SOLO_VENTA}
            GROUP BY EXTRACT(YEAR FROM v.fecha)
            ORDER BY anio ASC;
        """)

        query_zonas = text(f"""
            SELECT
                COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS zona,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades
            FROM db_ventas v
            WHERE v.fecha >= MAKE_DATE(:start_year, 1, 1)
              AND CAST(v.fecha AS DATE) <= MAKE_DATE(:end_year, 12, 31)
              AND {FILTRO_SCOPE_ROL}
              AND {FILTRO_VENDEDOR_OPCIONAL}
              AND {FILTRO_SOLO_VENTA}
            GROUP BY COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA')
            ORDER BY total_ventas DESC;
        """)

        query_top_clientes = text(f"""
            SELECT
                COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO') AS cliente,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades
            FROM db_ventas v
            WHERE v.fecha >= MAKE_DATE(:start_year, 1, 1)
              AND CAST(v.fecha AS DATE) <= MAKE_DATE(:end_year, 12, 31)
              AND {FILTRO_SCOPE_ROL}
              AND {FILTRO_VENDEDOR_OPCIONAL}
              AND {FILTRO_SOLO_VENTA}
            GROUP BY COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO')
            ORDER BY total_ventas DESC
            LIMIT 50;
        """)

        resumen_anual = [dict(r) for r in db.session.execute(query_resumen, params).mappings().all()]
        resumen_zonas = [dict(r) for r in db.session.execute(query_zonas, params).mappings().all()]
        top_clientes = [dict(r) for r in db.session.execute(query_top_clientes, params).mappings().all()]

        return {
            'success': True,
            'periodo': {'start_year': start_year, 'end_year': end_year},
            'seguridad': {'vista_global': es_global, 'usuario': username},
            'resumen_anual': resumen_anual,
            'resumen_zonas': resumen_zonas,
            'top_clientes': top_clientes
        }

    @staticmethod
    def obtener_detalle_paginado(user_id: int, username: str, user_role: str, start_year: int, end_year: int,
                                  pagina: int = 1, tam_pagina: int = 100, busqueda: str = '') -> dict:
        """
        Detalle consolidado (Año, Mes, Zona, Cliente) con paginación server-side.
        La agregación (GROUP BY) y el filtro de búsqueda corren en PostgreSQL;
        el LIMIT/OFFSET evita que el DTO crezca sin techo a medida que se acumulan años.
        """
        role_upper = str(user_role or '').strip().upper()
        es_global = role_upper in ['ADMIN', 'ADMINISTRACION', 'ADMINISTRADOR', 'GERENCIA']
        es_comercial = not es_global

        pagina = max(1, int(pagina or 1))
        tam_pagina = min(200, max(10, int(tam_pagina or 100)))  # techo duro: nunca mas de 200 filas por respuesta
        offset = (pagina - 1) * tam_pagina

        ventas_expr = ComercialHistoricoService._expr_ajustado_nc('total_ingresos')
        cantidad_expr = ComercialHistoricoService._expr_ajustado_nc('cantidad')
        vendedor_scope = ComercialHistoricoService._resolver_alias_vendedor(username)

        params = {
            'start_year': int(start_year),
            'end_year': int(end_year),
            'es_comercial': es_comercial,
            'vendedor_user': vendedor_scope,
            'vendedor_filtro': '',
            'busqueda': str(busqueda or '').strip(),
            'tam_pagina': tam_pagina,
            'offset': offset
        }

        query = text(f"""
            WITH agrupado AS (
                SELECT
                    EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                    EXTRACT(MONTH FROM v.fecha)::INTEGER AS mes,
                    COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS zona,
                    COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO') AS cliente,
                    ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                    ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                    COUNT(v.id)::INTEGER AS total_transacciones
                FROM db_ventas v
                WHERE v.fecha >= MAKE_DATE(:start_year, 1, 1)
                  AND CAST(v.fecha AS DATE) <= MAKE_DATE(:end_year, 12, 31)
                  AND {FILTRO_SCOPE_ROL}
                  AND {FILTRO_VENDEDOR_OPCIONAL}
                  AND {FILTRO_SOLO_VENTA}
                GROUP BY 1, 2, 3, 4
            )
            SELECT *, COUNT(*) OVER()::INTEGER AS total_registros
            FROM agrupado
            WHERE (:busqueda = '' OR cliente ILIKE '%' || :busqueda || '%' OR zona ILIKE '%' || :busqueda || '%')
            ORDER BY anio ASC, mes ASC, total_ventas DESC
            LIMIT :tam_pagina OFFSET :offset;
        """)

        try:
            filas = [dict(r) for r in db.session.execute(query, params).mappings().all()]
        except Exception as e:
            logger.error(f"[COMERCIAL_SERVICE] Error en detalle paginado: {e}")
            raise e

        total_registros = filas[0]['total_registros'] if filas else 0
        for f in filas:
            f.pop('total_registros', None)

        total_paginas = max(1, (total_registros + tam_pagina - 1) // tam_pagina)

        return {
            'success': True,
            'filas': filas,
            'paginacion': {
                'pagina': pagina,
                'tam_pagina': tam_pagina,
                'total_registros': total_registros,
                'total_paginas': total_paginas
            }
        }

    @staticmethod
    def obtener_crecimiento_clientes(user_id: int, username: str, user_role: str,
                                      anio_base: int, anio_comparacion: int, zona: str = None,
                                      busqueda: str = '', aplicar_ytd: bool = False) -> dict:
        """
        Compara ventas por cliente entre anio_base y anio_comparacion (YoY). La
        agregación por cliente/ciudad y el filtro de búsqueda por nombre corren en
        PostgreSQL; el cálculo de crecimiento (incluida la división por cero cuando
        el cliente es nuevo) se resuelve aquí, nunca en la ruta ni en el frontend.
        Respeta el mismo aislamiento por rol que el resto del servicio (FILTRO_SCOPE_ROL).

        aplicar_ytd=True corta ambos años en el mismo (mes, día) de hoy, para no
        comparar un año cerrado (12 meses) contra uno en curso (ej. 7 meses) — la
        falacia estadística que producía crecimientos negativos irreales cuando
        anio_comparacion es el año actual. Mismo criterio que _query_ytd, pero vía
        TO_CHAR(v.fecha, 'MM-DD') en vez de la tupla (mes, día).

        El corte YTD ahora vive dentro de cada CASE (no en el WHERE): además de
        venta_base/venta_comparacion (recortadas a YTD), también se necesita
        venta_base_anio_completo (SIN recortar) para poder distinguir un cliente
        que no compró en absoluto en el año base de uno que sí compró, pero después
        del corte YTD — son casos económicamente distintos ('REACTIVADO' vs
        'SIN_VENTA_EN_VENTANA') y solo se pueden separar teniendo ambas sumas.

        `clasificacion` resuelve la falacia de "NUEVO": antes, cualquier cliente con
        venta_base_ytd=0 se marcaba NUEVO, incluso si llevaba años comprando (ver
        caso HUGO ARMANDO BALLESTAS JOLY: su única compra en el año base fue en
        agosto, después del corte YTD de julio, y aun así aparecía "NUEVO"). Ahora
        se cruza con `primera_compra_historica` (MIN(fecha) de TODA la vida del
        cliente en db_ventas, sin importar año) para diferenciar:
          - NUEVO: su primera compra en la historia es del año de comparación en
            adelante (nunca compró antes).
          - REACTIVADO: tiene historia anterior al año base, pero no compró nada
            dentro del año base (ni siquiera fuera de la ventana YTD).
          - SIN_VENTA_EN_VENTANA: sí compró en el año base, pero después del corte
            YTD (por eso venta_base_ytd da 0 aunque el cliente esté activo).
          - ACTIVO: tiene venta_base_ytd > 0, se calcula el % de crecimiento normal.
        """
        role_upper = str(user_role or '').strip().upper()
        es_global = role_upper in ['ADMIN', 'ADMINISTRACION', 'ADMINISTRADOR', 'GERENCIA']
        es_comercial = not es_global

        ventas_expr = ComercialHistoricoService._expr_ajustado_nc('total_ingresos')
        vendedor_scope = ComercialHistoricoService._resolver_alias_vendedor(username)

        zona_normalizada = str(zona or '').strip()
        ciudades_zona = MAPEO_ZONAS.get(zona_normalizada, []) if zona_normalizada else []
        zona_activa = bool(zona_normalizada) and bool(ciudades_zona)
        if zona_normalizada and not ciudades_zona:
            logger.warning(f"[COMERCIAL_SERVICE] Zona '{zona_normalizada}' no está mapeada en MAPEO_ZONAS; se ignora el filtro.")

        limite_mmdd = datetime.now().strftime('%m-%d')

        params = {
            'anio_base': int(anio_base),
            'anio_comparacion': int(anio_comparacion),
            'es_comercial': es_comercial,
            'vendedor_user': vendedor_scope,
            'vendedor_filtro': '',
            'zona_activa': zona_activa,
            'ciudades_zona': ciudades_zona or [''],  # placeholder inocuo cuando zona_activa=False
            'busqueda': str(busqueda or '').strip(),
            'aplicar_ytd': bool(aplicar_ytd),
            'limite_mmdd': limite_mmdd,
        }

        # (:aplicar_ytd = FALSE OR TO_CHAR(...) <= :limite_mmdd) va DENTRO de cada
        # CASE (no en el WHERE): venta_base_anio_completo necesita las mismas filas
        # sin el recorte YTD para poder distinguir REACTIVADO de SIN_VENTA_EN_VENTANA.
        query = text(f"""
            WITH cliente_origen AS (
                -- Primera compra de CADA cliente en TODA la historia de db_ventas,
                -- sin filtrar por año/zona/búsqueda (solo scope de rol + "venta real"):
                -- filtrar por zona aquí sería incorrecto si el cliente cambió de
                -- ciudad entre su primera compra y ahora.
                SELECT
                    UPPER(TRIM(v.nombres)) AS cliente,
                    MIN(v.fecha) AS primera_compra_historica
                FROM db_ventas v
                WHERE {FILTRO_SCOPE_ROL}
                  AND {FILTRO_VENDEDOR_OPCIONAL}
                  AND {FILTRO_SOLO_VENTA}
                GROUP BY 1
            ),
            ventas_periodo AS (
                SELECT
                    COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO') AS cliente,
                    COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS ciudad,
                    ROUND(SUM(CASE WHEN EXTRACT(YEAR FROM v.fecha) = :anio_base
                                    AND (:aplicar_ytd = FALSE OR TO_CHAR(v.fecha, 'MM-DD') <= :limite_mmdd)
                                   THEN {ventas_expr} ELSE 0 END)::NUMERIC, 2) AS venta_base,
                    ROUND(SUM(CASE WHEN EXTRACT(YEAR FROM v.fecha) = :anio_comparacion
                                    AND (:aplicar_ytd = FALSE OR TO_CHAR(v.fecha, 'MM-DD') <= :limite_mmdd)
                                   THEN {ventas_expr} ELSE 0 END)::NUMERIC, 2) AS venta_comparacion,
                    ROUND(SUM(CASE WHEN EXTRACT(YEAR FROM v.fecha) = :anio_base
                                   THEN {ventas_expr} ELSE 0 END)::NUMERIC, 2) AS venta_base_anio_completo
                FROM db_ventas v
                WHERE EXTRACT(YEAR FROM v.fecha)::INTEGER IN (:anio_base, :anio_comparacion)
                  AND {FILTRO_SCOPE_ROL}
                  AND {FILTRO_VENDEDOR_OPCIONAL}
                  AND {FILTRO_SOLO_VENTA}
                  AND (:zona_activa = FALSE OR TRIM(UPPER(COALESCE(v.zona, ''))) = ANY(:ciudades_zona))
                  AND (:busqueda = '' OR v.nombres ILIKE '%' || :busqueda || '%')
                GROUP BY 1, 2
                HAVING SUM(CASE WHEN EXTRACT(YEAR FROM v.fecha) = :anio_base
                                 AND (:aplicar_ytd = FALSE OR TO_CHAR(v.fecha, 'MM-DD') <= :limite_mmdd)
                                THEN {ventas_expr} ELSE 0 END) <> 0
                    OR SUM(CASE WHEN EXTRACT(YEAR FROM v.fecha) = :anio_comparacion
                                 AND (:aplicar_ytd = FALSE OR TO_CHAR(v.fecha, 'MM-DD') <= :limite_mmdd)
                                THEN {ventas_expr} ELSE 0 END) <> 0
            )
            SELECT
                vp.cliente,
                vp.ciudad,
                vp.venta_base,
                vp.venta_comparacion,
                vp.venta_base_anio_completo,
                co.primera_compra_historica
            FROM ventas_periodo vp
            LEFT JOIN cliente_origen co ON co.cliente = UPPER(TRIM(vp.cliente));
        """)

        try:
            filas = [dict(r) for r in db.session.execute(query, params).mappings().all()]
        except Exception as e:
            logger.error(f"[COMERCIAL_SERVICE] Error generando crecimiento de clientes: {e}")
            raise e

        # Consolida por (cliente, zona-región): un mismo cliente puede aparecer bajo
        # más de una ciudad entre los dos años (traslado, sucursal); se suman antes
        # de calcular el crecimiento para no duplicarlo en la tabla final.
        # primera_compra_historica se agrupa solo por nombre de cliente (no por
        # ciudad) en la CTE, así que su valor ya es idéntico en todas las filas del
        # mismo cliente — basta con tomarlo una vez, no sumarlo.
        consolidado = {}
        for f in filas:
            ciudad_upper = str(f['ciudad']).strip().upper()
            zona_cliente = CIUDAD_A_ZONA.get(ciudad_upper, 'SIN ZONA')
            clave = (f['cliente'], zona_cliente)
            acc = consolidado.setdefault(clave, {
                'venta_base': 0.0, 'venta_comparacion': 0.0,
                'venta_base_anio_completo': 0.0, 'primera_compra_historica': f['primera_compra_historica']
            })
            acc['venta_base'] += float(f['venta_base'] or 0)
            acc['venta_comparacion'] += float(f['venta_comparacion'] or 0)
            acc['venta_base_anio_completo'] += float(f['venta_base_anio_completo'] or 0)

        # Lookup de NIT *después* de agregar las ventas (nunca antes vía JOIN):
        # db_ventas no tiene columna NIT, solo db_ventas.nombres (texto libre) y
        # db_clientes.identificacion; un JOIN previo al SUM podría duplicar ventas
        # si un nombre matchea más de un registro en db_clientes. DISTINCT ON
        # garantiza como máximo un NIT por nombre normalizado (el de menor id).
        nombres_normalizados = list({str(cliente).strip().upper() for (cliente, _z) in consolidado.keys()})
        nit_por_nombre = {}
        if nombres_normalizados:
            try:
                query_nits = text("""
                    SELECT DISTINCT ON (UPPER(TRIM(nombre)))
                        UPPER(TRIM(nombre)) AS nombre_normalizado,
                        identificacion AS nit
                    FROM db_clientes
                    WHERE UPPER(TRIM(nombre)) = ANY(:nombres_clientes)
                    ORDER BY UPPER(TRIM(nombre)), id ASC;
                """)
                filas_nits = db.session.execute(query_nits, {'nombres_clientes': nombres_normalizados}).mappings().all()
                nit_por_nombre = {r['nombre_normalizado']: r['nit'] for r in filas_nits}
            except Exception as e:
                logger.warning(f"[COMERCIAL_SERVICE] No se pudo resolver NIT de clientes: {e}")

        resultado = []
        for (cliente, zona_cliente), valores in consolidado.items():
            venta_base = round(valores['venta_base'], 2)
            venta_comparacion = round(valores['venta_comparacion'], 2)
            venta_base_anio_completo = round(valores['venta_base_anio_completo'], 2)
            primera_compra = valores['primera_compra_historica']

            # Sin match en cliente_origen (nombre en blanco -> 'CLIENTE DESCONOCIDO'
            # no cruza contra UPPER(TRIM(nombres))='' ): no hay forma de saber su
            # antigüedad real, se trata como NUEVO por defecto (no se puede probar
            # lo contrario).
            anio_primera_compra = primera_compra.year if primera_compra else anio_comparacion

            if anio_primera_compra >= anio_comparacion:
                clasificacion = 'NUEVO'
                es_nuevo = True
            elif venta_base == 0 and anio_primera_compra < anio_base:
                clasificacion = 'REACTIVADO'
                es_nuevo = False
            elif venta_base == 0 and venta_base_anio_completo > 0:
                clasificacion = 'SIN_VENTA_EN_VENTANA'
                es_nuevo = False
            elif venta_base > 0:
                clasificacion = 'ACTIVO'
                es_nuevo = False
            else:
                # Remanente no cubierto por las 4 reglas del Arquitecto: primera
                # compra cae en un año intermedio (anio_base < año < anio_comparacion,
                # solo posible si los años elegidos no son consecutivos) y venta_base=0
                # en todo el año base. Es un cliente que ya existía antes de
                # anio_comparacion, así que se trata igual que REACTIVADO.
                clasificacion = 'REACTIVADO'
                es_nuevo = False

            if clasificacion == 'ACTIVO':
                crecimiento_pct = round(((venta_comparacion - venta_base) / venta_base) * 100, 2)
            else:
                crecimiento_pct = 0.0

            resultado.append({
                'cliente_nit': nit_por_nombre.get(str(cliente).strip().upper()),
                'cliente_nombre': cliente,
                'zona_region': zona_cliente,
                'venta_anio_base': venta_base,
                'venta_anio_comp': venta_comparacion,
                'variacion_cop': round(venta_comparacion - venta_base, 2),
                'crecimiento_pct': crecimiento_pct,
                'clasificacion': clasificacion,
                'es_nuevo': es_nuevo
            })

        resultado.sort(key=lambda r: r['variacion_cop'], reverse=True)

        return {
            'success': True,
            'periodo': {
                'anio_base': int(anio_base),
                'anio_comparacion': int(anio_comparacion),
                'ytd_aplicado': bool(aplicar_ytd),
                'limite_mmdd': limite_mmdd if aplicar_ytd else None
            },
            'zona_filtro': zona_normalizada or None,
            'seguridad': {'vista_global': es_global, 'usuario': username},
            'clientes': resultado
        }

    @staticmethod
    def _formatear_hoja(worksheet, df, columnas_dinero=(), columnas_enteras=(), ancho_col=18):
        """
        Aplica formato profesional a una hoja ya escrita por pandas (index=False):
        encabezado con fondo corporativo, freeze panes, formato contable en columnas
        de dinero/unidades y ancho de columna legible. Reutilizable entre hojas.
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        fondo_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        fuente_header = Font(color='FFFFFF', bold=True)
        num_filas = len(df)

        for idx, nombre_col in enumerate(df.columns, start=1):
            letra = get_column_letter(idx)
            celda = worksheet.cell(row=1, column=idx)
            celda.fill = fondo_header
            celda.font = fuente_header
            celda.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.column_dimensions[letra].width = ancho_col

            if nombre_col in columnas_dinero:
                formato = '"$" #,##0'
            elif nombre_col in columnas_enteras:
                formato = '#,##0'
            else:
                formato = None

            if formato:
                for fila in range(2, num_filas + 2):
                    worksheet.cell(row=fila, column=idx).number_format = formato

        worksheet.freeze_panes = 'A2'

    @staticmethod
    def _formatear_hoja_pivot(worksheet, pivot_df, ancho_col=18):
        """
        Variante para hojas escritas con index=True (la pivot Y-o-Y): la primera
        columna es el índice (zona, texto) y el resto son años con valores de dinero.
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        fondo_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        fuente_header = Font(color='FFFFFF', bold=True)
        num_filas = len(pivot_df)
        num_columnas = len(pivot_df.columns) + 1  # +1 por la columna de índice (zona)

        for idx in range(1, num_columnas + 1):
            letra = get_column_letter(idx)
            celda = worksheet.cell(row=1, column=idx)
            celda.fill = fondo_header
            celda.font = fuente_header
            celda.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.column_dimensions[letra].width = ancho_col

            if idx > 1:  # columnas de años = dinero; columna 1 = nombre de zona (texto)
                for fila in range(2, num_filas + 2):
                    worksheet.cell(row=fila, column=idx).number_format = '"$" #,##0'

        worksheet.freeze_panes = 'B2'  # columna de zona siempre visible al hacer scroll horizontal

    @staticmethod
    def _formatear_hoja_comparativo_cliente(worksheet, pivot_df):
        """
        Formato para la hoja comparativa Cliente x Mes (replica el informe de World Office):
        columna 1 = cliente (texto), columnas intermedias = meses (dinero), última columna
        'Total' y última fila 'TOTAL VENTAS' resaltadas en negrita, igual que el reporte de WO.
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        fondo_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        fuente_header = Font(color='FFFFFF', bold=True)
        fondo_total = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        fuente_negrita = Font(bold=True)

        # pivot_df ya trae agregadas la columna 'Total' (derecha) y la fila 'TOTAL VENTAS' (abajo)
        num_filas = len(pivot_df)
        num_columnas = len(pivot_df.columns) + 1  # +1 por la columna de índice (cliente)
        fila_total_excel = num_filas + 1  # fila 1 = encabezado; ultima fila de datos = TOTAL VENTAS

        for idx in range(1, num_columnas + 1):
            letra = get_column_letter(idx)
            celda = worksheet.cell(row=1, column=idx)
            celda.fill = fondo_header
            celda.font = fuente_header
            celda.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.column_dimensions[letra].width = 32 if idx == 1 else 16

            for fila in range(2, fila_total_excel + 1):
                celda_dato = worksheet.cell(row=fila, column=idx)
                if idx > 1:
                    celda_dato.number_format = '"$" #,##0'
                if idx == num_columnas or fila == fila_total_excel:  # columna Total o fila TOTAL VENTAS
                    celda_dato.font = fuente_negrita
                    celda_dato.fill = fondo_total

        worksheet.freeze_panes = 'B2'

    @staticmethod
    def _query_ytd(user_id: int, username: str, user_role: str, start_year: int, end_year: int,
                    vendedor_filtro: str, corte_dt: date):
        """
        Ejecuta las 4 consultas de agregación YTD (anual, zona, mensual por zona y
        top clientes) 100% en SQL, todas bajo el mismo corte YTD y scope de vendedor
        para que las hojas del Excel sean comparables entre sí.
        El corte YTD compara (mes, dia) por tupla para ser exacto entre años
        bisiestos y no bisiestos (evita el corrimiento que introduce DOY tras el 29-feb).
        """
        role_upper = str(user_role or '').strip().upper()
        es_global = role_upper in ['ADMIN', 'ADMINISTRACION', 'ADMINISTRADOR', 'GERENCIA']
        es_comercial = not es_global

        ventas_expr = ComercialHistoricoService._expr_ajustado_nc('total_ingresos')
        cantidad_expr = ComercialHistoricoService._expr_ajustado_nc('cantidad')
        vendedor_scope = ComercialHistoricoService._resolver_alias_vendedor(username)

        params = {
            'start_year': int(start_year),
            'end_year': int(end_year),
            'es_comercial': es_comercial,
            'vendedor_user': vendedor_scope,
            'vendedor_filtro': str(vendedor_filtro or '').strip(),
            'corte_dt': corte_dt
        }

        where_ytd = f"""
            WHERE EXTRACT(YEAR FROM v.fecha)::INTEGER BETWEEN :start_year AND :end_year
              AND (EXTRACT(MONTH FROM v.fecha), EXTRACT(DAY FROM v.fecha))
                  <= (EXTRACT(MONTH FROM :corte_dt), EXTRACT(DAY FROM :corte_dt))
              AND {FILTRO_SCOPE_ROL}
              AND {FILTRO_VENDEDOR_OPCIONAL}
              AND {FILTRO_SOLO_VENTA}
        """

        query_anual = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                COUNT(v.id)::INTEGER AS total_transacciones
            FROM db_ventas v
            {where_ytd}
            GROUP BY EXTRACT(YEAR FROM v.fecha)
            ORDER BY anio ASC;
        """)

        query_zona = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS zona,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                COUNT(v.id)::INTEGER AS total_transacciones
            FROM db_ventas v
            {where_ytd}
            GROUP BY 1, 2
            ORDER BY zona ASC, anio ASC;
        """)

        # Desglose mensual por zona: profundidad temporal dentro de cada año
        query_mensual_zona = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                EXTRACT(MONTH FROM v.fecha)::INTEGER AS mes,
                COALESCE(NULLIF(TRIM(v.zona), ''), 'SIN ZONA') AS zona,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades
            FROM db_ventas v
            {where_ytd}
            GROUP BY 1, 2, 3
            ORDER BY anio ASC, mes ASC, zona ASC;
        """)

        # Top clientes del periodo/scope seleccionado (agregado, nunca transacciones sueltas)
        query_top_clientes = text(f"""
            SELECT
                COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO') AS cliente,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas,
                ROUND(SUM({cantidad_expr})::NUMERIC, 2) AS total_unidades,
                COUNT(v.id)::INTEGER AS total_transacciones
            FROM db_ventas v
            {where_ytd}
            GROUP BY 1
            ORDER BY total_ventas DESC
            LIMIT 30;
        """)

        # Ventas mensuales por cliente (comparativo, replica estructura del informe de World Office).
        # Sin LIMIT: a diferencia de una tabla en el navegador, un Excel con cientos de filas no
        # tiene costo de renderizado — respeta el mismo scope que el resto de hojas.
        query_mensual_cliente = text(f"""
            SELECT
                EXTRACT(YEAR FROM v.fecha)::INTEGER AS anio,
                EXTRACT(MONTH FROM v.fecha)::INTEGER AS mes,
                COALESCE(NULLIF(TRIM(v.nombres), ''), 'CLIENTE DESCONOCIDO') AS cliente,
                ROUND(SUM({ventas_expr})::NUMERIC, 2) AS total_ventas
            FROM db_ventas v
            {where_ytd}
            GROUP BY 1, 2, 3
            ORDER BY cliente ASC, anio ASC, mes ASC;
        """)

        try:
            ytd_anual = [dict(r) for r in db.session.execute(query_anual, params).mappings().all()]
            ytd_zona = [dict(r) for r in db.session.execute(query_zona, params).mappings().all()]
            ytd_mensual_zona = [dict(r) for r in db.session.execute(query_mensual_zona, params).mappings().all()]
            top_clientes = [dict(r) for r in db.session.execute(query_top_clientes, params).mappings().all()]
            ytd_mensual_cliente = [dict(r) for r in db.session.execute(query_mensual_cliente, params).mappings().all()]
        except Exception as e:
            logger.error(f"[COMERCIAL_SERVICE] Error generando agregación YTD: {e}")
            raise e

        return ytd_anual, ytd_zona, ytd_mensual_zona, top_clientes, ytd_mensual_cliente, es_global

    @staticmethod
    def generar_excel_ytd_stream(user_id: int, username: str, user_role: str, start_year: int, end_year: int,
                                  vendedor_filtro: str = '', fecha_corte=None):
        """
        Genera el .xlsx de analítica comercial YTD/Y-o-Y con formato profesional
        (encabezado corporativo, freeze panes, formato contable) y devuelve
        (buffer, nombre_archivo) listo para pasarle a send_file. Encapsula toda la
        dependencia de pandas/openpyxl para que el controlador quede limpio.
        """
        import pandas as pd
        import io

        corte_dt = ComercialHistoricoService._resolver_fecha_corte(fecha_corte)

        ytd_anual, ytd_zona, ytd_mensual_zona, top_clientes, ytd_mensual_cliente, _ = ComercialHistoricoService._query_ytd(
            user_id=user_id, username=username, user_role=user_role,
            start_year=start_year, end_year=end_year,
            vendedor_filtro=vendedor_filtro, corte_dt=corte_dt
        )

        df_anual = pd.DataFrame(ytd_anual)
        df_zona = pd.DataFrame(ytd_zona)
        df_mensual_zona = pd.DataFrame(ytd_mensual_zona)
        df_top_clientes = pd.DataFrame(top_clientes)
        df_mensual_cliente = pd.DataFrame(ytd_mensual_cliente)

        # Reshape puro (no reagrega nada): pivotea filas ya sumadas por SQL para lectura Y-o-Y.
        if not df_zona.empty:
            pivot_yoy = df_zona.pivot_table(index='zona', columns='anio', values='total_ventas', fill_value=0)
        else:
            pivot_yoy = pd.DataFrame()

        # Comparativo Cliente x Mes (replica el informe "Comparativo Agrupado Por Vendedor" de WO):
        # reshape puro de filas ya agregadas por SQL, más fila/columna de totales (sumas, no recálculo).
        MESES_NOMBRE = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
            7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        if not df_mensual_cliente.empty:
            df_mensual_cliente = df_mensual_cliente.copy()
            df_mensual_cliente['periodo'] = df_mensual_cliente.apply(
                lambda r: f"{MESES_NOMBRE[int(r['mes'])]} {int(r['anio'])}", axis=1
            )
            orden_periodos = (
                df_mensual_cliente.drop_duplicates(subset=['anio', 'mes'])
                .sort_values(['anio', 'mes'])['periodo'].tolist()
            )
            pivot_cliente_mes = df_mensual_cliente.pivot_table(
                index='cliente', columns='periodo', values='total_ventas', aggfunc='sum', fill_value=0
            )[orden_periodos]
            pivot_cliente_mes['Total'] = pivot_cliente_mes.sum(axis=1)
            fila_total = pivot_cliente_mes.sum(axis=0)
            fila_total.name = 'TOTAL VENTAS'
            pivot_cliente_mes = pd.concat([pivot_cliente_mes, fila_total.to_frame().T])
        else:
            pivot_cliente_mes = pd.DataFrame()

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_anual.to_excel(writer, index=False, sheet_name='YTD Anual')
            ComercialHistoricoService._formatear_hoja(
                writer.sheets['YTD Anual'], df_anual,
                columnas_dinero=('total_ventas',), columnas_enteras=('total_unidades', 'total_transacciones')
            )

            df_zona.to_excel(writer, index=False, sheet_name='YTD por Zona')
            ComercialHistoricoService._formatear_hoja(
                writer.sheets['YTD por Zona'], df_zona,
                columnas_dinero=('total_ventas',), columnas_enteras=('total_unidades', 'total_transacciones')
            )

            df_mensual_zona.to_excel(writer, index=False, sheet_name='Desglose Mensual Zona')
            ComercialHistoricoService._formatear_hoja(
                writer.sheets['Desglose Mensual Zona'], df_mensual_zona,
                columnas_dinero=('total_ventas',), columnas_enteras=('total_unidades',)
            )

            df_top_clientes.to_excel(writer, index=False, sheet_name='Top Clientes')
            ComercialHistoricoService._formatear_hoja(
                writer.sheets['Top Clientes'], df_top_clientes,
                columnas_dinero=('total_ventas',), columnas_enteras=('total_unidades', 'total_transacciones'),
                ancho_col=32  # nombres de cliente suelen ser largos
            )

            pivot_yoy.to_excel(writer, sheet_name='YoY Zona (Pivot)')
            if not pivot_yoy.empty:
                ComercialHistoricoService._formatear_hoja_pivot(writer.sheets['YoY Zona (Pivot)'], pivot_yoy)

            pivot_cliente_mes.to_excel(writer, sheet_name='Ventas Mensuales Cliente')
            if not pivot_cliente_mes.empty:
                ComercialHistoricoService._formatear_hoja_comparativo_cliente(
                    writer.sheets['Ventas Mensuales Cliente'], pivot_cliente_mes
                )
        output.seek(0)

        vendedor_slug = re.sub(r'[^A-Za-z0-9_-]+', '_', vendedor_filtro.strip())[:40] if vendedor_filtro else 'GLOBAL'
        nombre_archivo = f"Comercial_YTD_{start_year}-{end_year}_{vendedor_slug}_{corte_dt.isoformat()}.xlsx"

        return output, nombre_archivo
